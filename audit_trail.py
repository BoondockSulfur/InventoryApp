"""
Enhanced Audit Trail System
- Detailed change history with diffs
- Export audit logs to various formats
- Advanced filtering and search
- Diff view for changes
"""

import json
import csv
import io
from datetime import datetime
from difflib import unified_diff
from flask import Blueprint, jsonify, request, send_file
from flask_login import login_required, current_user

audit_bp = Blueprint('audit', __name__, url_prefix='/audit')


class AuditTrailService:
    """Service for managing audit logs"""

    def __init__(self, db):
        self.db = db

    def log_change(self, entity_type, entity_id, action, changes, user=None, ip_address=None):
        """
        Log a change to an entity

        Args:
            entity_type: Type of entity (item, loan, ticket, etc.)
            entity_id: ID of the entity
            action: Action performed (create, update, delete)
            changes: Dict of changes {field: {'old': old_value, 'new': new_value}}
            user: User who made the change
            ip_address: IP address of the request

        Returns:
            AuditLog object
        """
        from app import AuditLog

        try:
            audit_log = AuditLog(
                entity_type=entity_type,
                entity_id=entity_id,
                action=action,
                changes=json.dumps(changes),
                user_id=user.id if user else None,
                ip_address=ip_address,
                timestamp=datetime.utcnow()
            )

            self.db.session.add(audit_log)
            self.db.session.commit()

            return audit_log
        except Exception as e:
            self.db.session.rollback()
            raise e

    def get_entity_history(self, entity_type, entity_id):
        """Get complete history of an entity"""
        from app import AuditLog, User

        logs = AuditLog.query.filter_by(
            entity_type=entity_type,
            entity_id=entity_id
        ).order_by(AuditLog.timestamp.desc()).all()

        history = []
        for log in logs:
            changes = json.loads(log.changes) if log.changes else {}

            history.append({
                'id': log.id,
                'action': log.action,
                'user': log.user.username if log.user else 'System',
                'timestamp': log.timestamp.isoformat(),
                'changes': changes,
                'ip_address': log.ip_address
            })

        return history

    def get_diff_view(self, entity_type, entity_id, log_id=None):
        """
        Get a diff view of changes

        Args:
            entity_type: Type of entity
            entity_id: ID of the entity
            log_id: Specific log ID to show diff for (or latest if None)

        Returns:
            Dict with diff information
        """
        from app import AuditLog

        if log_id:
            log = AuditLog.query.get(log_id)
        else:
            log = AuditLog.query.filter_by(
                entity_type=entity_type,
                entity_id=entity_id
            ).order_by(AuditLog.timestamp.desc()).first()

        if not log:
            return None

        changes = json.loads(log.changes) if log.changes else {}

        # Create diff for each changed field
        diffs = {}
        for field, change_data in changes.items():
            old_value = str(change_data.get('old', ''))
            new_value = str(change_data.get('new', ''))

            if old_value != new_value:
                diff = list(unified_diff(
                    old_value.splitlines(keepends=True),
                    new_value.splitlines(keepends=True),
                    fromfile=f'{field} (old)',
                    tofile=f'{field} (new)',
                    lineterm=''
                ))
                diffs[field] = {
                    'old': old_value,
                    'new': new_value,
                    'diff': diff
                }

        return {
            'log_id': log.id,
            'timestamp': log.timestamp.isoformat(),
            'user': log.user.username if log.user else 'System',
            'action': log.action,
            'diffs': diffs
        }

    def export_audit_logs(self, format='csv', filters=None):
        """
        Export audit logs to various formats

        Args:
            format: Export format (csv, json, xlsx)
            filters: Dict of filters to apply

        Returns:
            File-like object with exported data
        """
        from app import AuditLog, User

        # Build query with filters
        query = AuditLog.query

        if filters:
            if filters.get('entity_type'):
                query = query.filter_by(entity_type=filters['entity_type'])
            if filters.get('entity_id'):
                query = query.filter_by(entity_id=filters['entity_id'])
            if filters.get('action'):
                query = query.filter_by(action=filters['action'])
            if filters.get('user_id'):
                query = query.filter_by(user_id=filters['user_id'])
            if filters.get('date_from'):
                query = query.filter(AuditLog.timestamp >= filters['date_from'])
            if filters.get('date_to'):
                query = query.filter(AuditLog.timestamp <= filters['date_to'])

        logs = query.order_by(AuditLog.timestamp.desc()).all()

        if format == 'csv':
            return self._export_to_csv(logs)
        elif format == 'json':
            return self._export_to_json(logs)
        elif format == 'xlsx':
            return self._export_to_xlsx(logs)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _export_to_csv(self, logs):
        """Export logs to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow(['ID', 'Timestamp', 'Entity Type', 'Entity ID', 'Action', 'User', 'IP Address', 'Changes'])

        # Write data
        for log in logs:
            changes = json.loads(log.changes) if log.changes else {}
            changes_str = '; '.join([f"{k}: {v['old']} → {v['new']}" for k, v in changes.items()])

            writer.writerow([
                log.id,
                log.timestamp.isoformat(),
                log.entity_type,
                log.entity_id,
                log.action,
                log.user.username if log.user else 'System',
                log.ip_address or '',
                changes_str
            ])

        output.seek(0)
        return output

    def _export_to_json(self, logs):
        """Export logs to JSON format"""
        data = []

        for log in logs:
            changes = json.loads(log.changes) if log.changes else {}

            data.append({
                'id': log.id,
                'timestamp': log.timestamp.isoformat(),
                'entity_type': log.entity_type,
                'entity_id': log.entity_id,
                'action': log.action,
                'user': log.user.username if log.user else 'System',
                'user_id': log.user_id,
                'ip_address': log.ip_address,
                'changes': changes
            })

        output = io.StringIO()
        json.dump(data, output, indent=2)
        output.seek(0)
        return output

    def _export_to_xlsx(self, logs):
        """Export logs to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Audit Logs"

            # Header with styling
            headers = ['ID', 'Timestamp', 'Entity Type', 'Entity ID', 'Action', 'User', 'IP Address', 'Changes']
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_idx, log in enumerate(logs, 2):
                changes = json.loads(log.changes) if log.changes else {}
                changes_str = '\n'.join([f"{k}: {v.get('old', '')} → {v.get('new', '')}" for k, v in changes.items()])

                ws.cell(row=row_idx, column=1, value=log.id)
                ws.cell(row=row_idx, column=2, value=log.timestamp.isoformat())
                ws.cell(row=row_idx, column=3, value=log.entity_type)
                ws.cell(row=row_idx, column=4, value=log.entity_id)
                ws.cell(row=row_idx, column=5, value=log.action)
                ws.cell(row=row_idx, column=6, value=log.user.username if log.user else 'System')
                ws.cell(row=row_idx, column=7, value=log.ip_address or '')
                ws.cell(row=row_idx, column=8, value=changes_str)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        except ImportError:
            raise Exception("openpyxl library required for Excel export")

    def get_user_activity_summary(self, user_id, days=30):
        """Get activity summary for a user"""
        from app import AuditLog
        from datetime import timedelta

        start_date = datetime.utcnow() - timedelta(days=days)

        logs = AuditLog.query.filter(
            AuditLog.user_id == user_id,
            AuditLog.timestamp >= start_date
        ).all()

        summary = {
            'total_actions': len(logs),
            'actions_by_type': {},
            'entities_modified': {},
            'daily_activity': {}
        }

        for log in logs:
            # Count by action type
            action = log.action
            summary['actions_by_type'][action] = summary['actions_by_type'].get(action, 0) + 1

            # Count by entity type
            entity_type = log.entity_type
            summary['entities_modified'][entity_type] = summary['entities_modified'].get(entity_type, 0) + 1

            # Count by day
            day = log.timestamp.date().isoformat()
            summary['daily_activity'][day] = summary['daily_activity'].get(day, 0) + 1

        return summary


# ============================================================================
# API Routes
# ============================================================================

@audit_bp.route('/api/logs', methods=['GET'])
@login_required
def get_audit_logs():
    """Get audit logs with filtering"""
    from app import AuditLog, db

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    entity_type = request.args.get('entity_type')
    entity_id = request.args.get('entity_id', type=int)
    action = request.args.get('action')
    user_id = request.args.get('user_id', type=int)

    query = AuditLog.query

    if entity_type:
        query = query.filter_by(entity_type=entity_type)
    if entity_id:
        query = query.filter_by(entity_id=entity_id)
    if action:
        query = query.filter_by(action=action)
    if user_id:
        query = query.filter_by(user_id=user_id)

    pagination = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    logs = [{
        'id': log.id,
        'entity_type': log.entity_type,
        'entity_id': log.entity_id,
        'action': log.action,
        'user': log.user.username if log.user else 'System',
        'timestamp': log.timestamp.isoformat(),
        'ip_address': log.ip_address,
        'changes': json.loads(log.changes) if log.changes else {}
    } for log in pagination.items]

    return jsonify({
        'success': True,
        'logs': logs,
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': pagination.total,
            'pages': pagination.pages
        }
    })


@audit_bp.route('/api/logs/export', methods=['GET'])
@login_required
def export_logs():
    """Export audit logs"""
    from app import db

    format = request.args.get('format', 'csv')
    filters = {
        'entity_type': request.args.get('entity_type'),
        'entity_id': request.args.get('entity_id', type=int),
        'action': request.args.get('action'),
        'user_id': request.args.get('user_id', type=int),
    }

    service = AuditTrailService(db)
    output = service.export_audit_logs(format=format, filters=filters)

    if format == 'csv':
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'audit_logs_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    elif format == 'json':
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='application/json',
            as_attachment=True,
            download_name=f'audit_logs_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.json'
        )
    elif format == 'xlsx':
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'audit_logs_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )


@audit_bp.route('/api/logs/<int:log_id>/diff', methods=['GET'])
@login_required
def get_log_diff(log_id):
    """Get diff view for a specific log"""
    from app import db

    service = AuditTrailService(db)
    log = service.db.session.get(service.db.Model, log_id)

    if not log:
        return jsonify({'success': False, 'error': 'Log not found'}), 404

    diff = service.get_diff_view(log.entity_type, log.entity_id, log_id)

    return jsonify({
        'success': True,
        'diff': diff
    })


@audit_bp.route('/api/entity/<entity_type>/<int:entity_id>/history', methods=['GET'])
@login_required
def get_entity_history(entity_type, entity_id):
    """Get complete history of an entity"""
    from app import db

    service = AuditTrailService(db)
    history = service.get_entity_history(entity_type, entity_id)

    return jsonify({
        'success': True,
        'history': history
    })


# Global audit service instance (will be initialized with db in app.py)
audit_service = None
