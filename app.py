# -*- coding: utf-8 -*-

APP_VERSION = "2.3.0"  # Full Implementation: i18n, Analytics, WebSocket, GraphQL, Reporting, Integrations, Mobile API

import os
import uuid
import socket
import csv, subprocess
import logging
import requests
from logging.handlers import RotatingFileHandler
from io import BytesIO
from datetime import date, datetime, timedelta
from flask import (
    Flask, send_file, render_template, redirect, url_for,
    flash, request, jsonify
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, login_user, login_required,
    logout_user, current_user, UserMixin
)
from flask_mail import Mail, Message
from flask_wtf import FlaskForm, CSRFProtect
from wtforms import (
    StringField, PasswordField, SubmitField,
    SelectField, DateField, EmailField,
    IntegerField, BooleanField, TextAreaField, DecimalField
)
from wtforms.validators import (
    DataRequired, Length, Email, EqualTo, ValidationError, Regexp, Optional
)
from werkzeug.security import generate_password_hash, check_password_hash
from apscheduler.schedulers.background import BackgroundScheduler

# WeasyPrint optional (benötigt System-Libraries auf Windows)
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    WEASYPRINT_AVAILABLE = False
    print(f"⚠️  WeasyPrint not available: {e}")
    print("   PDF generation will be disabled. This is OK for local testing.")

from dotenv import load_dotenv
from functools import wraps
from flask import abort
from flask import current_app, g
from itsdangerous import URLSafeTimedSerializer as Serializer
from sqlalchemy import or_, func, desc, and_
from sqlalchemy.orm import joinedload, selectinload
import qrcode
from io import BytesIO
import base64
from PIL import Image
import json
from werkzeug.utils import secure_filename
import hashlib
from difflib import SequenceMatcher

# Optional imports - features degrade gracefully if not available
try:
    from icalendar import Calendar, Event as iCalEvent
    import pytz
    ICALENDAR_AVAILABLE = True
except ImportError:
    ICALENDAR_AVAILABLE = False
    print("⚠️  icalendar not available - iCal export will be disabled")

try:
    import pyotp
    PYOTP_AVAILABLE = True
except ImportError:
    PYOTP_AVAILABLE = False
    print("⚠️  pyotp not available - 2FA will be disabled")

try:
    from authlib.integrations.flask_client import OAuth
    OAUTH_AVAILABLE = True
except ImportError:
    OAUTH_AVAILABLE = False
    print("⚠️  authlib not available - OAuth login will be disabled")

try:
    from onelogin.saml2.auth import OneLogin_Saml2_Auth
    from onelogin.saml2.utils import OneLogin_Saml2_Utils
    SAML_AVAILABLE = True
except ImportError:
    SAML_AVAILABLE = False
    print("⚠️  python3-saml not available - SAML SSO will be disabled")

try:
    from flask_babel import Babel, gettext, lazy_gettext as _l
    from flask import session
    BABEL_AVAILABLE = True
except ImportError:
    BABEL_AVAILABLE = False
    print("⚠️  Flask-Babel not available - Multi-language support will be disabled")
    # Fallback: gettext returns string unchanged
    def gettext(s): return s
    _l = gettext

try:
    from flask_socketio import SocketIO, emit, join_room, leave_room
    SOCKETIO_AVAILABLE = True
except ImportError:
    SOCKETIO_AVAILABLE = False
    print("⚠️  Flask-SocketIO not available - Real-time updates will be disabled")

try:
    import graphene
    from flask_graphql import GraphQLView
    GRAPHQL_AVAILABLE = True
except ImportError:
    GRAPHQL_AVAILABLE = False
    print("⚠️  GraphQL libraries not available - GraphQL API will be disabled")

try:
    import plotly
    import plotly.graph_objs as go
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    print("⚠️  Plotly not available - Advanced charts will be disabled")

from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# NEW FEATURES: Import new services
from elasticsearch_service import es_service
from ml_service import ml_service
from audit_trail import AuditTrailService, audit_bp
from mobile_api import mobile_api
from search_api import search_api
from sqlalchemy import event

# ─── Import Utils: Constants & Helpers ────────────────────────────────────────
from utils.constants import (
    UserRole, ItemStatus, TicketStatus, TicketPriorityLevel,
    TicketCategory as TicketCategoryConstants,
    MaintenanceStatus, HTTPStatus, FlashCategory, AuditAction, Permission
)
from utils.auth_helpers import (
    flash_success, flash_error, flash_warning, flash_info,
    safe_commit, safe_add, safe_delete,
    check_user_permission, check_role_permission,
    log_user_action, log_security_event
)
from utils.validators import (
    validate_email, validate_serial_number, sanitize_filename,
    validate_url, validate_barcode, validate_rfid_tag
)

# ─── Umgebungsvariablen laden ─────────────────────────────────────────────────
load_dotenv()

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY')
    if not SECRET_KEY:
        raise ValueError("SECRET_KEY must be set in environment variables")

    SQLALCHEMY_DATABASE_URI = os.environ.get('DATABASE_URL', 'sqlite:///inventory.db')
    SQLALCHEMY_TRACK_MODIFICATIONS = False

    # Security Settings
    SESSION_COOKIE_SECURE = os.environ.get('SESSION_COOKIE_SECURE', 'true').lower() in ['true', '1', 'yes']
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = 'Lax'
    PERMANENT_SESSION_LIFETIME = 3600
    MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # 16 MB file upload limit

    MAIL_SERVER   = os.environ.get('MAIL_SERVER')
    MAIL_PORT     = int(os.environ.get('MAIL_PORT', 587))
    MAIL_USE_TLS  = os.environ.get('MAIL_USE_TLS', 'false').lower() in ['true','1','yes']
    MAIL_USERNAME = os.environ.get('MAIL_USERNAME')
    MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD')
    ADMINS        = os.environ.get('ADMIN_EMAIL','').split(',') if os.environ.get('ADMIN_EMAIL') else []

    PRINTER_IP   = os.environ.get('PRINTER_IP')
    PRINTER_PORT = int(os.environ.get('PRINTER_PORT', 9100))
    WARRANTY_CSV_DIR = os.environ.get('WARRANTY_CSV_DIR')
    WARRANTY_CLI_CMD = os.environ.get('WARRANTY_CLI_CMD', 'DellWarranty-CLI.exe')
    DELL_CLIENT_ID     = os.environ.get('DELL_CLIENT_ID')
    DELL_CLIENT_SECRET = os.environ.get('DELL_CLIENT_SECRET')

    # Babel i18n Configuration
    BABEL_DEFAULT_LOCALE = 'de'
    BABEL_DEFAULT_TIMEZONE = 'Europe/Berlin'
    LANGUAGES = {'de': 'Deutsch', 'en': 'English'}

    # SocketIO Configuration
    SOCKETIO_MESSAGE_QUEUE = os.environ.get('REDIS_URL', 'redis://redis:6379/0')
    SOCKETIO_ASYNC_MODE = 'threading'

# ─── App-Initialisierung ──────────────────────────────────────────────────────
app = Flask(__name__)
app.config.from_object(Config)
if not app.debug:
    log_path = app.config.get(
        'ERROR_LOG_PATH',
        os.path.join(app.instance_path, 'error.log')
    )
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    fh = RotatingFileHandler(log_path, maxBytes=10*1024*1024, backupCount=3)
    fh.setLevel(logging.ERROR)
    app.logger.addHandler(fh)
app.jinja_env.globals['date'] = date
csrf  = CSRFProtect(app)
db    = SQLAlchemy(app)
login = LoginManager(app)
login.login_view = 'auth.login'

# NEW FEATURES: Initialize services
es_service.init_app(app)
ml_service.init_app(app)
audit_service = AuditTrailService(db)
app.register_blueprint(audit_bp)
app.register_blueprint(mobile_api)
app.register_blueprint(search_api)
app.logger.info('[OK] New features initialized')

role_permissions = db.Table(
    'role_permissions',
    db.Column('role_name',       db.String(20), db.ForeignKey('role.name'),       primary_key=True),
    db.Column('permission_name', db.String(50), db.ForeignKey('permission.name'), primary_key=True)
)

user_groups = db.Table(
    'user_groups',
    db.Column('user_id',  db.Integer,    db.ForeignKey('user.id'),       primary_key=True),
    db.Column('group_id', db.Integer,    db.ForeignKey('group.id'),      primary_key=True)
)

group_roles = db.Table(
    'group_roles',
    db.Column('group_id',  db.Integer,   db.ForeignKey('group.id'),      primary_key=True),
    db.Column('role_name', db.String(20), db.ForeignKey('role.name'),    primary_key=True)
)

# User loader für Flask-Login
@login.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

mail      = Mail(app)

# Babel i18n Initialization
babel = None
if BABEL_AVAILABLE:
    from flask_babel import Babel
    babel = Babel()

    def get_locale():
        # 1. Try URL parameter
        lang = request.args.get('lang')
        if lang in app.config['LANGUAGES']:
            session['language'] = lang
            return lang
        # 2. Try session
        if 'language' in session:
            return session['language']
        # 3. Try user preference from database
        if current_user.is_authenticated:
            user_lang = db.session.execute(
                db.select(ModuleSetting).filter_by(
                    key='user_language',
                    user_id=current_user.id
                )
            ).scalar_one_or_none()
            if user_lang and user_lang.value in app.config['LANGUAGES']:
                return user_lang.value
        # 4. Try browser language
        return request.accept_languages.best_match(app.config['LANGUAGES'].keys())

    babel.init_app(app, locale_selector=get_locale)

# SocketIO Initialization
socketio = None
if SOCKETIO_AVAILABLE:
    socketio = SocketIO(
        app,
        message_queue=app.config['SOCKETIO_MESSAGE_QUEUE'],
        async_mode=app.config['SOCKETIO_ASYNC_MODE'],
        cors_allowed_origins="*"
    )

# OAuth Initialization (optional)
oauth = None
if OAUTH_AVAILABLE:
    oauth = OAuth(app)

    # Google OAuth
    if os.environ.get('GOOGLE_CLIENT_ID') and os.environ.get('GOOGLE_CLIENT_SECRET'):
        oauth.register(
            name='google',
            client_id=os.environ.get('GOOGLE_CLIENT_ID'),
            client_secret=os.environ.get('GOOGLE_CLIENT_SECRET'),
            server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
            client_kwargs={'scope': 'openid email profile'}
        )

    # Microsoft OAuth
    if os.environ.get('MICROSOFT_CLIENT_ID') and os.environ.get('MICROSOFT_CLIENT_SECRET'):
        oauth.register(
            name='microsoft',
            client_id=os.environ.get('MICROSOFT_CLIENT_ID'),
            client_secret=os.environ.get('MICROSOFT_CLIENT_SECRET'),
            server_metadata_url='https://login.microsoftonline.com/common/v2.0/.well-known/openid-configuration',
            client_kwargs={'scope': 'openid email profile'}
        )

    # GitHub OAuth
    if os.environ.get('GITHUB_CLIENT_ID') and os.environ.get('GITHUB_CLIENT_SECRET'):
        oauth.register(
            name='github',
            client_id=os.environ.get('GITHUB_CLIENT_ID'),
            client_secret=os.environ.get('GITHUB_CLIENT_SECRET'),
            access_token_url='https://github.com/login/oauth/access_token',
            access_token_params=None,
            authorize_url='https://github.com/login/oauth/authorize',
            authorize_params=None,
            api_base_url='https://api.github.com/',
            client_kwargs={'scope': 'user:email'}
        )

scheduler = BackgroundScheduler()

# Rate Limiting - Use Redis if available, otherwise fallback to memory
redis_url = os.environ.get('REDIS_URL')
storage_uri = redis_url if redis_url else "memory://"

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri=storage_uri,
    strategy="fixed-window"
)

# ─── Datenbank-Modelle ────────────────────────────────────────────────────────
class Role(db.Model):
    __tablename__ = 'role'
    name        = db.Column(db.String(20), primary_key=True)
    permissions = db.relationship(
        'Permission',
        secondary=role_permissions,
        back_populates='roles'
    )

class Permission(db.Model):
    __tablename__ = 'permission'
    name        = db.Column(db.String(50), primary_key=True)
    description = db.Column(db.String(200))
    roles       = db.relationship(
        'Role',
        secondary=role_permissions,
        back_populates='permissions'
    )

class Group(db.Model):
    __tablename__ = 'group'
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(100), unique=True, nullable=False, index=True)
    description = db.Column(db.String(500))
    created_at  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    users = db.relationship(
        'User',
        secondary=user_groups,
        back_populates='groups'
    )
    roles = db.relationship(
        'Role',
        secondary=group_roles,
        backref=db.backref('groups', lazy='dynamic')
    )

class Setting(db.Model):
    """
    Schlüssel–Wert Paare für konfigurierbare System-Einstellungen.
    """
    key   = db.Column(db.String(50), primary_key=True)
    value = db.Column(db.String(200), nullable=False)

class Company(db.Model):
    """
    Multi-Mandanten-Fähigkeit: Organisationen/Firmen
    """
    __tablename__ = 'company'
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(100), unique=True, nullable=False, index=True)
    subdomain   = db.Column(db.String(50), unique=True, nullable=True, index=True)  # Optional für URL-basierte Trennung
    logo_url    = db.Column(db.String(256), nullable=True)
    primary_color = db.Column(db.String(7), default='#0d6efd')  # Hex color for branding

    # Contact & Address
    address     = db.Column(db.String(256), nullable=True)
    phone       = db.Column(db.String(50), nullable=True)
    email       = db.Column(db.String(120), nullable=True)
    website     = db.Column(db.String(256), nullable=True)

    # Settings
    is_active   = db.Column(db.Boolean, default=True, nullable=False)
    max_users   = db.Column(db.Integer, nullable=True)  # Optional user limit
    max_items   = db.Column(db.Integer, nullable=True)  # Optional item limit

    created_at  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    users = db.relationship('User', backref='company', lazy='dynamic')
    items = db.relationship('Item', backref='company', lazy='dynamic')

class User(UserMixin, db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(64), unique=True, nullable=False, index=True)
    email         = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(128), nullable=False)
    role          = db.Column(db.String(20), nullable=False, default='verwaltung', index=True)
    company_id    = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=True, index=True)  # Multi-Mandanten
    group_number  = db.Column(db.String(32), nullable=True)
    email_verified = db.Column(db.Boolean, default=False)
    verification_token = db.Column(db.String(128), unique=True, nullable=True)
    totp_secret    = db.Column(db.String(32), nullable=True)  # 2FA Secret
    totp_enabled   = db.Column(db.Boolean, default=False)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Feature 18: Dark Mode
    theme_preference = db.Column(db.String(20), default='auto')  # 'light', 'dark', 'auto'

    # Feature 16: Slack/Teams Integration
    slack_webhook = db.Column(db.String(256), nullable=True)
    teams_webhook = db.Column(db.String(256), nullable=True)

    # Feature 17: LDAP Integration
    ldap_dn = db.Column(db.String(256), nullable=True)
    is_ldap_user = db.Column(db.Boolean, default=False)

    # Notification Preferences
    email_notifications = db.Column(db.Boolean, default=True)
    push_notifications = db.Column(db.Boolean, default=True)
    digest_frequency = db.Column(db.String(20), default='weekly')  # 'daily', 'weekly', 'never'

    # Group Membership
    groups = db.relationship(
        'Group',
        secondary=user_groups,
        back_populates='users'
    )

    def set_password(self, pw):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw):
        return check_password_hash(self.password_hash, pw)

    def get_reset_token(self, expires_sec=3600):
        s = Serializer(app.config['SECRET_KEY'], expires_sec)
        return s.dumps({'user_id': self.id})

    @staticmethod
    def verify_reset_token(token):
        s = Serializer(app.config['SECRET_KEY'])
        try:
            data = s.loads(token)
        except (Exception, ValueError):
            return None
        return User.query.get(data['user_id'])

    def has_permission(self, permission_name):
        """
        Prüft ob der Benutzer eine bestimmte Berechtigung hat.
        Admins haben immer alle Berechtigungen.
        Berücksichtigt Rollen über direkte Rolle und Gruppen.
        """
        # Admin hat immer alle Rechte
        if self.role == UserRole.ADMIN:
            return True

        # Sammle alle Permissions aus der direkten Rolle
        permissions = set()
        role = db.session.get(Role, self.role)
        if role:
            permissions.update(p.name for p in role.permissions)

        # Sammle alle Permissions aus Gruppen
        for group in self.groups:
            for group_role in group.roles:
                permissions.update(p.name for p in group_role.permissions)

        return permission_name in permissions

    def has_any_permission(self, *permission_names):
        """Prüft ob der Benutzer mindestens eine der angegebenen Berechtigungen hat."""
        return any(self.has_permission(perm) for perm in permission_names)

    def has_all_permissions(self, *permission_names):
        """Prüft ob der Benutzer alle angegebenen Berechtigungen hat."""
        return all(self.has_permission(perm) for perm in permission_names)

    def get_all_permissions(self):
        """Gibt alle Berechtigungen des Benutzers zurück."""
        if self.role == UserRole.ADMIN:
            return set(p.name for p in Permission.query.all())

        permissions = set()
        role = db.session.get(Role, self.role)
        if role:
            permissions.update(p.name for p in role.permissions)

        for group in self.groups:
            for group_role in group.roles:
                permissions.update(p.name for p in group_role.permissions)

        return permissions

    @property
    def is_admin(self):
        """Schnelle Prüfung ob Benutzer Admin ist."""
        return self.role == UserRole.ADMIN

    @property
    def is_manager(self):
        """Prüft ob Benutzer Verwaltungsrechte hat."""
        return self.role in ['admin', 'verwaltung'] or self.has_permission('manage_items')

class Item(db.Model):
    """Core inventory item model with comprehensive tracking features.

    This model represents individual items in the inventory system with support for:
    - Multi-tenant company assignments
    - QR codes, barcodes, and RFID tags
    - Budget and value tracking with depreciation
    - Location management and item hierarchy
    - Team assignments and maintenance scheduling
    - Dell warranty/contract integration

    Attributes:
        id: Primary key identifier
        name: Item display name
        serial: Unique serial number
        company_id: Company assignment for multi-tenant support
        location: Legacy location string
        location_id: Foreign key to Location model
        category_id: Foreign key to Category model
        team_id: Foreign key to Team model
        parent_item_id: Self-referential FK for item hierarchy
    """
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(128), nullable=False, index=True)
    serial      = db.Column(db.String(64), unique=True, nullable=False, index=True)
    company_id  = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=True, index=True)  # Multi-Mandanten
    location    = db.Column(db.String(128))
    note        = db.Column(db.Text)
    category    = db.Column(db.String(32), nullable=False, default='other', index=True)
    is_borrowed = db.Column(db.Boolean, default=False, index=True)
    defective   = db.Column(db.Boolean, default=False, index=True)
    image_path  = db.Column(db.String(256), nullable=True)  # Pfad zum hochgeladenen Bild
    created_at  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at  = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Feature 1: QR-Code System
    qr_code = db.Column(db.Text, nullable=True)  # Base64-encoded QR-Code

    # Feature 23: Barcode Scanner
    barcode = db.Column(db.String(128), unique=True, nullable=True, index=True)

    # Feature 24: RFID Integration
    rfid_tag_id = db.Column(db.String(64), unique=True, nullable=True, index=True)

    # Feature 7: Budget & Value Tracking
    purchase_price = db.Column(db.Float, nullable=True)
    purchase_date = db.Column(db.Date, nullable=True)
    depreciation_rate = db.Column(db.Float, default=0.0)  # Prozent pro Jahr
    current_value = db.Column(db.Float, nullable=True)

    # Feature 8: Location Management
    location_id = db.Column(db.Integer, db.ForeignKey('location.id', ondelete='SET NULL'), nullable=True, index=True)

    # NEW: Category as Foreign Key (flexible user-created categories)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id', ondelete='SET NULL'), nullable=True, index=True)

    # NEW: Item Hierarchy (items can belong to other items, like laptop in workstation)
    parent_item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='SET NULL'), nullable=True, index=True)

    # Feature 14: Team Management
    team_id = db.Column(db.Integer, db.ForeignKey('team.id', ondelete='SET NULL'), nullable=True, index=True)

    # Feature 13: Approval Workflow
    requires_approval = db.Column(db.Boolean, default=False)

    # Feature 12: Rating System
    average_rating = db.Column(db.Float, default=0.0)
    rating_count = db.Column(db.Integer, default=0)

    # Feature 2: Maintenance Tracking
    next_maintenance_date = db.Column(db.Date, nullable=True, index=True)
    maintenance_interval_days = db.Column(db.Integer, nullable=True)  # Wartungsintervall in Tagen

    # Dell Warranty/Contract Integration
    dell_service_tag = db.Column(db.String(64), nullable=True, index=True)  # Dell Service Tag
    dell_warranty_status = db.Column(db.String(50), nullable=True)  # Active, Expired, etc.
    dell_warranty_end_date = db.Column(db.Date, nullable=True, index=True)
    dell_contract_type = db.Column(db.String(100), nullable=True)  # ProSupport, Basic, etc.
    dell_last_sync = db.Column(db.DateTime, nullable=True)  # Letzter Sync-Zeitpunkt

    # NEW Relationships
    category_rel = db.relationship('Category', backref='items', foreign_keys=[category_id])
    parent_item = db.relationship('Item', remote_side=[id], backref='child_items', foreign_keys=[parent_item_id])
    # Note: location_rel uses existing backref 'location_detail' from Location model
    # overlaps parameter resolves the conflict between location_rel and location_detail
    location_rel = db.relationship('Location', foreign_keys=[location_id], overlaps="location_detail")

class Loan(db.Model):
    """Item loan tracking model.

    Tracks when items are borrowed by users, including loan dates,
    due dates, and return dates. Supports active loan status checking.

    Attributes:
        id: Primary key identifier
        item_id: Foreign key to borrowed Item
        borrower_id: Foreign key to User who borrowed the item
        loan_date: Date item was loaned out
        due_date: Date item is due to be returned
        return_date: Actual return date (null if not yet returned)
    """
    __tablename__ = 'loan'
    id           = db.Column(db.Integer, primary_key=True)
    item_id      = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    borrower_id  = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    loan_date    = db.Column(db.Date, default=date.today, nullable=False, index=True)
    due_date     = db.Column(db.Date, nullable=False, index=True)
    return_date  = db.Column(db.Date, index=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    borrower     = db.relationship('User', backref='loans')
    item         = db.relationship('Item', backref='loans')

class Ticket(db.Model):
    __tablename__ = 'ticket'
    id           = db.Column(db.Integer, primary_key=True)
    title        = db.Column(db.String(150), nullable=False)
    description  = db.Column(db.Text, nullable=False)
    status       = db.Column(db.String(20), nullable=False, default='open')  # open, in_progress, waiting, resolved, closed
    priority     = db.Column(db.String(20), nullable=False, default='medium')  # low, medium, high, urgent
    category     = db.Column(db.String(50), nullable=True)  # hardware, software, access, other
    assigned_to  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Zugewiesener Supporter
    created_at   = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    resolved_at  = db.Column(db.DateTime, nullable=True)
    closed_at    = db.Column(db.DateTime, nullable=True)
    due_date     = db.Column(db.DateTime, nullable=True)  # SLA-Frist
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

    # Relationships
    user         = db.relationship('User', foreign_keys=[user_id], backref='tickets')
    assignee     = db.relationship('User', foreign_keys=[assigned_to], backref='assigned_tickets')
    responses    = db.relationship('TicketResponse', backref='ticket', cascade='all, delete-orphan', order_by='TicketResponse.created_at')

    # Metadata
    tags         = db.Column(db.String(200), nullable=True)  # Comma-separated tags
    satisfaction_rating = db.Column(db.Integer, nullable=True)  # 1-5 stars

class TicketResponse(db.Model):
    __tablename__ = 'ticket_response'
    id           = db.Column(db.Integer, primary_key=True)
    ticket_id    = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    user         = db.relationship('User', backref='ticket_responses')
    message      = db.Column(db.Text, nullable=False)
    is_internal  = db.Column(db.Boolean, default=False)  # Interne Notiz (nicht für User sichtbar)
    is_solution  = db.Column(db.Boolean, default=False)  # Als Lösung markiert
    created_at   = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)
    attachments  = db.Column(db.String(500), nullable=True)  # JSON mit Datei-Pfaden

# ============================================================================
# NEW FEATURES: Elasticsearch Auto-Indexing Hooks
# ============================================================================

@event.listens_for(Item, 'after_insert')
def item_after_insert(mapper, connection, target):
    """Auto-index item in Elasticsearch after insert"""
    try:
        from flask import current_app
        current_app.logger.info(f'Indexing new item: {target.id}')
        es_service.index_item(target)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error indexing item {target.id}: {e}')


@event.listens_for(Item, 'after_update')
def item_after_update(mapper, connection, target):
    """Auto-index item in Elasticsearch after update"""
    try:
        from flask import current_app
        current_app.logger.info(f'Re-indexing updated item: {target.id}')
        es_service.index_item(target)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error re-indexing item {target.id}: {e}')


@event.listens_for(Item, 'after_delete')
def item_after_delete(mapper, connection, target):
    """Remove item from Elasticsearch after delete"""
    try:
        from flask import current_app
        current_app.logger.info(f'Removing item from index: {target.id}')
        es_service.delete_item(target.id)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error removing item {target.id}: {e}')


@event.listens_for(Ticket, 'after_insert')
def ticket_after_insert(mapper, connection, target):
    """Auto-index ticket in Elasticsearch after insert"""
    try:
        from flask import current_app
        current_app.logger.info(f'Indexing new ticket: {target.id}')
        es_service.index_ticket(target)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error indexing ticket {target.id}: {e}')


@event.listens_for(Ticket, 'after_update')
def ticket_after_update(mapper, connection, target):
    """Auto-index ticket in Elasticsearch after update"""
    try:
        from flask import current_app
        current_app.logger.info(f'Re-indexing updated ticket: {target.id}')
        es_service.index_ticket(target)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error re-indexing ticket {target.id}: {e}')


@event.listens_for(Ticket, 'after_delete')
def ticket_after_delete(mapper, connection, target):
    """Remove ticket from Elasticsearch after delete"""
    try:
        from flask import current_app
        current_app.logger.info(f'Removing ticket from index: {target.id}')
        es_service.delete_ticket(target.id)
    except Exception as e:
        from flask import current_app
        current_app.logger.error(f'Error removing ticket {target.id}: {e}')

class AuditLog(db.Model):
    __tablename__ = 'audit_log'
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='SET NULL'), nullable=True, index=True)
    user       = db.relationship('User', backref='audit_logs')
    action     = db.Column(db.String(100), nullable=False, index=True)  # 'create_item', 'update_item', 'delete_user', etc.
    entity_type = db.Column(db.String(50), nullable=False)  # 'item', 'user', 'loan', etc.
    entity_id  = db.Column(db.Integer, nullable=True)
    changes    = db.Column(db.Text, nullable=True)  # JSON mit Änderungen
    ip_address = db.Column(db.String(45), nullable=True)  # IPv6-ready
    endpoint   = db.Column(db.String(200), nullable=False, index=True)
    method     = db.Column(db.String(10), nullable=False, index=True)
    timestamp  = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)

# ─── NEUE MODELLE FÜR FEATURES ────────────────────────────────────────────────

# Feature 8: Hierarchisches Standort-Management
class Location(db.Model):
    """Location management model with hierarchical support.

    Represents physical locations where items can be stored. Supports
    hierarchical organization (locations within locations), floor plans,
    and coordinate-based positioning.

    Attributes:
        id: Primary key identifier
        name: Location name
        description: Detailed location description
        parent_id: Self-referential FK for location hierarchy
        floor_plan_path: Path to uploaded floor plan image
        coordinates: JSON string with x/y coordinates
    """
    __tablename__ = 'location'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False, index=True)
    description = db.Column(db.Text, nullable=True)
    parent_id = db.Column(db.Integer, db.ForeignKey('location.id', ondelete='CASCADE'), nullable=True, index=True)
    floor_plan_path = db.Column(db.String(256), nullable=True)  # Grundriss-Upload
    coordinates = db.Column(db.String(100), nullable=True)  # JSON: {"x": 100, "y": 200}
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    parent = db.relationship('Location', remote_side=[id], backref='children')
    # overlaps parameter resolves the conflict between location_detail and location_rel in Item model
    items = db.relationship('Item', backref='location_detail', foreign_keys='Item.location_id', overlaps="location_rel")

# Feature: Category Management (for v2.3.0 compatibility)
class Category(db.Model):
    """Flexible category management for inventory items.

    Allows users to create custom categories instead of using hardcoded ones.
    Each category can have a name, description, and optional icon/color.

    Attributes:
        id: Primary key identifier
        name: Unique category name
        description: Category description
        icon: Icon identifier for UI display
        color: Hex color code for category
    """
    __tablename__ = 'category'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False, unique=True, index=True)
    description = db.Column(db.Text, nullable=True)
    icon = db.Column(db.String(50), default='folder')
    color = db.Column(db.String(7), default='#6c757d')
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

# Feature: Module Settings (for v2.3.0 integration features)
class ModuleSetting(db.Model):
    __tablename__ = 'module_setting'
    id = db.Column(db.Integer, primary_key=True)
    module_name = db.Column(db.String(64), nullable=False, index=True)
    setting_key = db.Column(db.String(128), nullable=False, index=True)
    setting_value = db.Column(db.Text, nullable=True)
    setting_type = db.Column(db.String(32), default='string')  # string, int, bool, json
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    __table_args__ = (
        db.UniqueConstraint('module_name', 'setting_key', name='uq_module_setting'),
    )

# Feature 14: Team-Verwaltung
class Team(db.Model):
    """Model for team management with enhanced features.

    Stores team information including basic details, contact information,
    permissions, and status. Teams can be assigned items and have members.
    """
    __tablename__ = 'team'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False, unique=True, index=True)
    description = db.Column(db.Text, nullable=True)
    department = db.Column(db.String(128), nullable=True, index=True)
    team_lead = db.Column(db.String(128), nullable=True)  # Name of team leader
    email = db.Column(db.String(128), nullable=True)  # Team contact email
    phone = db.Column(db.String(20), nullable=True)  # Team contact phone

    # Permissions
    can_view_all = db.Column(db.Boolean, default=False)
    can_edit_all = db.Column(db.Boolean, default=False)
    can_delete_all = db.Column(db.Boolean, default=False)
    can_manage_team = db.Column(db.Boolean, default=False)

    # Status
    is_active = db.Column(db.Boolean, default=True, index=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    items = db.relationship('Item', backref='team', foreign_keys='Item.team_id')
    members = db.relationship('TeamMember', backref='team', cascade='all, delete-orphan')

class TeamMember(db.Model):
    """Association model for team membership.

    Links users to teams with role assignments. Tracks when users
    joined teams and their role within the team structure.

    Attributes:
        id: Primary key identifier
        team_id: Foreign key to Team
        user_id: Foreign key to User
        role: Member role ('lead', 'member', 'moderator', etc.)
        joined_at: Timestamp when user joined the team
    """
    __tablename__ = 'team_member'
    id = db.Column(db.Integer, primary_key=True)
    team_id = db.Column(db.Integer, db.ForeignKey('team.id', ondelete='CASCADE'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    role = db.Column(db.String(20), default='member')  # 'lead', 'member'
    joined_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Relationships
    user = db.relationship('User', backref='team_memberships')

# Feature 3: Reservierungssystem
class Reservation(db.Model):
    __tablename__ = 'reservation'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    start_date = db.Column(db.DateTime, nullable=False, index=True)
    end_date = db.Column(db.DateTime, nullable=False, index=True)
    status = db.Column(db.String(20), default='pending')  # 'pending', 'confirmed', 'cancelled', 'completed'
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    item = db.relationship('Item', backref='reservations')
    user = db.relationship('User', backref='reservations')

# Feature 2: Wartungs-Tracker
class Maintenance(db.Model):
    __tablename__ = 'maintenance'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    performed_by_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='SET NULL'), nullable=True, index=True)
    scheduled_date = db.Column(db.Date, nullable=False, index=True)
    completed_date = db.Column(db.Date, nullable=True)
    maintenance_type = db.Column(db.String(50), nullable=False)  # 'inspection', 'repair', 'calibration', 'cleaning'
    status = db.Column(db.String(20), default='scheduled')  # 'scheduled', 'overdue', 'completed', 'cancelled'
    notes = db.Column(db.Text, nullable=True)
    cost = db.Column(db.Float, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    item = db.relationship('Item', backref='maintenance_records')
    performed_by = db.relationship('User', backref='performed_maintenances')

# Feature 11: Defekt-Meldung und Reparatur-Workflow
class DefectReport(db.Model):
    __tablename__ = 'defect_report'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    reported_by_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='SET NULL'), nullable=True, index=True)
    description = db.Column(db.Text, nullable=False)
    severity = db.Column(db.String(20), default='medium')  # 'low', 'medium', 'high', 'critical'
    status = db.Column(db.String(20), default='reported')  # 'reported', 'acknowledged', 'in_repair', 'fixed', 'unfixable'
    image_path = db.Column(db.String(256), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    item = db.relationship('Item', backref='defect_reports')
    reported_by = db.relationship('User', backref='reported_defects')
    repair_logs = db.relationship('RepairLog', backref='defect', cascade='all, delete-orphan')

class RepairLog(db.Model):
    __tablename__ = 'repair_log'
    id = db.Column(db.Integer, primary_key=True)
    defect_report_id = db.Column(db.Integer, db.ForeignKey('defect_report.id', ondelete='CASCADE'), nullable=False, index=True)
    repaired_by_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='SET NULL'), nullable=True, index=True)
    action_taken = db.Column(db.Text, nullable=False)
    parts_replaced = db.Column(db.Text, nullable=True)
    cost = db.Column(db.Float, nullable=True)
    repair_date = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Relationships
    repaired_by = db.relationship('User', backref='repair_logs')

# Feature 12: Kommentar- und Rating-System
class Comment(db.Model):
    __tablename__ = 'comment'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    content = db.Column(db.Text, nullable=False)
    is_internal = db.Column(db.Boolean, default=False)  # Nur für Admins sichtbar
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    item = db.relationship('Item', backref='comments')
    user = db.relationship('User', backref='comments')

class Rating(db.Model):
    __tablename__ = 'rating'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    stars = db.Column(db.Integer, nullable=False)  # 1-5
    review = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Unique constraint: Ein User kann ein Item nur einmal bewerten
    __table_args__ = (db.UniqueConstraint('item_id', 'user_id', name='unique_item_user_rating'),)

    # Relationships
    item = db.relationship('Item', backref='ratings')
    user = db.relationship('User', backref='ratings')

# Feature 13: Genehmigungs-Workflow
class ApprovalRequest(db.Model):
    __tablename__ = 'approval_request'
    id = db.Column(db.Integer, primary_key=True)
    item_id = db.Column(db.Integer, db.ForeignKey('item.id', ondelete='CASCADE'), nullable=False, index=True)
    requester_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    approver_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='SET NULL'), nullable=True, index=True)
    reason = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='pending')  # 'pending', 'approved', 'rejected'
    rejection_reason = db.Column(db.Text, nullable=True)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    item = db.relationship('Item', backref='approval_requests')
    requester = db.relationship('User', foreign_keys=[requester_id], backref='requested_approvals')
    approver = db.relationship('User', foreign_keys=[approver_id], backref='approved_requests')

# Feature 5 & 6: Benachrichtigungssystem
class Notification(db.Model):
    __tablename__ = 'notification'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    type = db.Column(db.String(50), nullable=False)  # 'loan_overdue', 'reservation_confirmed', 'maintenance_due', etc.
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    link = db.Column(db.String(256), nullable=True)  # Link zur relevanten Seite
    is_read = db.Column(db.Boolean, default=False, index=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)

    # Relationships
    user = db.relationship('User', backref='notifications')

# Feature 19: Gespeicherte Filter
class SavedFilter(db.Model):
    __tablename__ = 'saved_filter'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=False, index=True)
    name = db.Column(db.String(100), nullable=False)
    filter_type = db.Column(db.String(50), nullable=False)  # 'items', 'loans', 'users', etc.
    filter_params = db.Column(db.Text, nullable=False)  # JSON mit Filterparametern
    is_default = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Relationships
    user = db.relationship('User', backref='saved_filters')

# Feature 22: Granulares Berechtigungssystem
class UserPermission(db.Model):
    __tablename__ = 'user_permission'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), nullable=True, index=True)
    role = db.Column(db.String(20), nullable=True)  # Falls rollenbasiert
    resource_type = db.Column(db.String(50), nullable=False)  # 'item', 'loan', 'user', 'team', etc.
    resource_id = db.Column(db.Integer, nullable=True)  # Spezifische Ressource (NULL = alle)
    action = db.Column(db.String(50), nullable=False)  # 'view', 'create', 'update', 'delete', 'approve', etc.
    granted = db.Column(db.Boolean, default=True)
    expires_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Relationships
    user = db.relationship('User', backref='user_permissions')

# Feature 7: Budget-Tracking
class BudgetEntry(db.Model):
    """Model for budget entries with enhanced tracking capabilities.

    Stores budget information including name, category, amount, date range,
    department assignment, and spending tracking.
    """
    __tablename__ = 'budget_entry'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(128), nullable=False, index=True)  # Budget name/description
    year = db.Column(db.Integer, nullable=False, index=True)
    category = db.Column(db.String(50), nullable=False, index=True)
    allocated_amount = db.Column(db.Float, nullable=False)
    spent_amount = db.Column(db.Float, default=0.0)
    start_date = db.Column(db.Date, nullable=True, index=True)  # Budget period start
    end_date = db.Column(db.Date, nullable=True, index=True)  # Budget period end
    description = db.Column(db.Text, nullable=True)  # Detailed description
    department = db.Column(db.String(50), nullable=True, index=True)  # Department assignment
    notes = db.Column(db.Text, nullable=True)  # Legacy notes field
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    # Relationships
    transactions = db.relationship('BudgetTransaction', backref='budget', cascade='all, delete-orphan', lazy='dynamic')

    def calculate_spent_amount(self):
        """Calculate total spent amount from all expense, purchase, and depreciation transactions."""
        total_expenses = db.session.query(func.sum(BudgetTransaction.amount)).filter(
            BudgetTransaction.budget_id == self.id,
            BudgetTransaction.transaction_type.in_(['expense', 'purchase', 'depreciation'])
        ).scalar() or 0.0

        total_income = db.session.query(func.sum(BudgetTransaction.amount)).filter(
            BudgetTransaction.budget_id == self.id,
            BudgetTransaction.transaction_type == 'income'
        ).scalar() or 0.0

        return total_expenses - total_income

    def get_remaining_budget(self):
        """Calculate remaining budget amount."""
        return self.allocated_amount - self.calculate_spent_amount()

class BudgetTransaction(db.Model):
    """Model for budget transactions (expenses, income, depreciation).

    Tracks individual transactions that affect budget entries, including
    purchases, expenses, income, and depreciation calculations.
    """
    __tablename__ = 'budget_transaction'
    id = db.Column(db.Integer, primary_key=True)
    budget_id = db.Column(db.Integer, db.ForeignKey('budget_entry.id', ondelete='CASCADE'), nullable=False, index=True)
    amount = db.Column(db.Float, nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False, index=True)  # expense, income, depreciation, purchase
    transaction_date = db.Column(db.Date, nullable=False, default=date.today, index=True)
    description = db.Column(db.Text, nullable=True)
    reference_type = db.Column(db.String(50), nullable=True)  # 'item', 'maintenance', 'repair', etc.
    reference_id = db.Column(db.Integer, nullable=True)  # ID of related item/maintenance/repair
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    # Relationships
    creator = db.relationship('User', backref='budget_transactions')

# ═══════════════════════════════════════════════════════════════════════════
# ─── ENHANCED TICKET SYSTEM MODELS ─────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════

class TicketCategory(db.Model):
    """Configurable ticket categories"""
    __tablename__ = 'ticket_category'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    description = db.Column(db.String(200))
    color = db.Column(db.String(7), default='#6c757d')  # Hex color
    icon = db.Column(db.String(50), default='tag')  # Font Awesome icon
    is_active = db.Column(db.Boolean, default=True)
    default_assignee_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    sla_hours = db.Column(db.Integer, default=24)  # Default SLA in hours
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    default_assignee = db.relationship('User', backref='default_categories')

class TicketPriority(db.Model):
    """Configurable ticket priorities"""
    __tablename__ = 'ticket_priority'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(20), nullable=False, unique=True)
    level = db.Column(db.Integer, nullable=False)  # 1=lowest, 5=highest
    color = db.Column(db.String(7), default='#6c757d')
    response_time_hours = db.Column(db.Integer, default=24)
    resolution_time_hours = db.Column(db.Integer, default=72)
    auto_escalate = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TicketStatus(db.Model):
    """Configurable ticket statuses with workflow"""
    __tablename__ = 'ticket_status'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False, unique=True)
    display_name = db.Column(db.String(50), nullable=False)
    description = db.Column(db.String(200))
    color = db.Column(db.String(7), default='#6c757d')
    icon = db.Column(db.String(50), default='circle')
    is_closed = db.Column(db.Boolean, default=False)  # Marks ticket as closed
    is_resolved = db.Column(db.Boolean, default=False)  # Marks ticket as resolved
    order = db.Column(db.Integer, default=0)  # Display order
    allowed_next_statuses = db.Column(db.Text)  # JSON list of next possible statuses
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TicketCustomField(db.Model):
    """Custom fields for tickets"""
    __tablename__ = 'ticket_custom_field'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    field_type = db.Column(db.String(20), nullable=False)  # text, number, date, select, checkbox
    options = db.Column(db.Text)  # JSON for select options
    is_required = db.Column(db.Boolean, default=False)
    default_value = db.Column(db.String(200))
    help_text = db.Column(db.String(200))
    order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class TicketFieldValue(db.Model):
    """Values for custom fields"""
    __tablename__ = 'ticket_field_value'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    field_id = db.Column(db.Integer, db.ForeignKey('ticket_custom_field.id'), nullable=False)
    value = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    ticket = db.relationship('Ticket', backref='custom_field_values')
    field = db.relationship('TicketCustomField', backref='values')

class TicketTemplate(db.Model):
    """Templates for creating tickets"""
    __tablename__ = 'ticket_template'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(200))
    title_template = db.Column(db.String(150), nullable=False)
    description_template = db.Column(db.Text, nullable=False)
    default_category = db.Column(db.String(50))
    default_priority = db.Column(db.String(20))
    default_assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    assignee = db.relationship('User', backref='ticket_templates')

class TicketSLA(db.Model):
    """SLA rules for tickets"""
    __tablename__ = 'ticket_sla'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(200))
    priority_level = db.Column(db.Integer)  # Apply to specific priority
    category_name = db.Column(db.String(50))  # Apply to specific category
    first_response_hours = db.Column(db.Integer, default=4)
    resolution_hours = db.Column(db.Integer, default=24)
    notify_before_breach_hours = db.Column(db.Integer, default=2)
    escalate_on_breach = db.Column(db.Boolean, default=False)
    escalate_to_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    escalate_to = db.relationship('User', backref='sla_escalations')

class TicketTimeEntry(db.Model):
    """Time tracking for tickets"""
    __tablename__ = 'ticket_time_entry'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    minutes = db.Column(db.Integer, nullable=False)
    description = db.Column(db.String(500))
    is_billable = db.Column(db.Boolean, default=True)
    hourly_rate = db.Column(db.Float, default=0.0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    ticket = db.relationship('Ticket', backref='time_entries')
    user = db.relationship('User', backref='time_entries')

class TicketLink(db.Model):
    """Links between tickets"""
    __tablename__ = 'ticket_link'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    linked_ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    link_type = db.Column(db.String(20), nullable=False)  # blocks, blocked_by, relates_to, duplicates
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

    # Relationships
    ticket = db.relationship('Ticket', foreign_keys=[ticket_id], backref='outgoing_links')
    linked_ticket = db.relationship('Ticket', foreign_keys=[linked_ticket_id], backref='incoming_links')
    created_by = db.relationship('User', backref='ticket_links')

class TicketWatcher(db.Model):
    """Users watching a ticket for notifications"""
    __tablename__ = 'ticket_watcher'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('ticket.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    notify_on_update = db.Column(db.Boolean, default=True)
    notify_on_response = db.Column(db.Boolean, default=True)
    notify_on_status_change = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    ticket = db.relationship('Ticket', backref='watchers')
    user = db.relationship('User', backref='watched_tickets')

# ─── Formulare ─────────────────────────────────────────────────────────────────
# ─── Import Forms from forms.py ───────────────────────────────────────────────
from forms import (
    DummyForm, LoginForm, ResetPasswordRequestForm, ResetPasswordForm,
    InitialSetupForm, BudgetForm, BudgetTransactionForm, TeamForm,
    RegistrationForm, ItemForm, CategoryForm, LocationForm,
    BorrowerForm, ReturnForm, TicketForm, ResponseForm, StatusForm,
    EditUserForm, SettingsForm
)

# ─── Hilfsfunktionen ──────────────────────────────────────────────────────────

# Feature 1: QR-Code Generation
def generate_qr_code(item_id, base_url):
    """Generiert QR-Code für ein Item"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    item_url = f"{base_url}/item/{item_id}"
    qr.add_data(item_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)

    # Als Base64 kodieren
    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return f"data:image/png;base64,{img_base64}"

# Feature 10: Smart Suggestions - Ähnlichkeitsberechnung
def calculate_similarity(str1, str2):
    """Berechnet Ähnlichkeit zwischen zwei Strings (0.0 - 1.0)"""
    return SequenceMatcher(None, str1.lower(), str2.lower()).ratio()

def find_similar_items(item_name, category=None, limit=5):
    """Findet ähnliche Items basierend auf Name und Kategorie"""
    query = Item.query.filter(Item.name != item_name)
    if category:
        query = query.filter_by(category=category)

    all_items = query.all()
    similarities = [(item, calculate_similarity(item_name, item.name)) for item in all_items]
    similarities.sort(key=lambda x: x[1], reverse=True)

    return [item for item, score in similarities[:limit] if score > 0.3]

# Feature 10: Auto-Kategorisierung
def auto_categorize_item(item_name):
    """Schlägt automatisch eine Kategorie basierend auf dem Namen vor"""
    name_lower = item_name.lower()

    category_keywords = {
        'laptop': ['laptop', 'notebook', 'thinkpad', 'macbook', 'chromebook'],
        'monitor': ['monitor', 'display', 'bildschirm', 'screen'],
        'tablet': ['tablet', 'ipad'],
        'phone': ['phone', 'smartphone', 'handy', 'iphone'],
        'camera': ['kamera', 'camera', 'canon', 'nikon'],
        'projector': ['beamer', 'projektor', 'projector'],
        'printer': ['drucker', 'printer'],
        'tool': ['werkzeug', 'tool', 'bohrer', 'säge', 'hammer'],
        'furniture': ['möbel', 'stuhl', 'tisch', 'schrank', 'chair', 'desk'],
    }

    for category, keywords in category_keywords.items():
        if any(keyword in name_lower for keyword in keywords):
            return category

    return 'other'

# Feature 7: Wertberechnung mit Abschreibung
def calculate_current_value(purchase_price, purchase_date, depreciation_rate):
    """Berechnet aktuellen Wert unter Berücksichtigung der Abschreibung"""
    if not purchase_price or not purchase_date:
        return None

    years_old = (date.today() - purchase_date).days / 365.25
    depreciation = purchase_price * (depreciation_rate / 100) * years_old
    current_value = max(0, purchase_price - depreciation)

    return round(current_value, 2)

# Feature 5: Benachrichtigungen erstellen
def create_notification(user_id, notification_type, title, message, link=None):
    """Erstellt eine neue Benachrichtigung für einen Benutzer"""
    notification = Notification(
        user_id=user_id,
        type=notification_type,
        title=title,
        message=message,
        link=link
    )
    db.session.add(notification)
    db.session.commit()

    # Optional: Push-Benachrichtigung senden
    user = User.query.get(user_id)
    if user and user.push_notifications:
        # TODO: Web Push API implementieren
        pass

    return notification

# Feature 16: Slack/Teams Benachrichtigung
def send_webhook_notification(webhook_url, title, message, color='info'):
    """Sendet Benachrichtigung an Slack oder Teams Webhook"""
    if not webhook_url:
        return False

    try:
        # Slack Format
        if 'slack.com' in webhook_url:
            payload = {
                "text": f"*{title}*\n{message}",
                "color": color
            }
        # Teams Format
        else:
            payload = {
                "@type": "MessageCard",
                "@context": "https://schema.org/extensions",
                "summary": title,
                "title": title,
                "text": message,
                "themeColor": "0076D7" if color == 'info' else "FF0000"
            }

        response = requests.post(webhook_url, json=payload, timeout=5)
        return response.status_code == 200
    except Exception as e:
        app.logger.error(f"Webhook notification failed: {e}")
        return False

# Feature 15: iCal Export
def create_ical_event(title, start_date, end_date, description='', location=''):
    """Erstellt ein iCal Event"""
    if not ICALENDAR_AVAILABLE:
        return None

    cal = Calendar()
    cal.add('prodid', '-//InventoryApp//inventoryapp//EN')
    cal.add('version', '2.0')

    event = iCalEvent()
    event.add('summary', title)
    event.add('dtstart', start_date)
    event.add('dtend', end_date)
    event.add('description', description)
    if location:
        event.add('location', location)

    cal.add_component(event)
    return cal.to_ical()

# Feature 21: Audit Log Helper
def log_action(action, entity_type, entity_id=None, changes=None):
    """Loggt eine Aktion im Audit Log"""
    try:
        audit_entry = AuditLog(
            user_id=current_user.id if current_user.is_authenticated else None,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            changes=json.dumps(changes) if changes else None,
            ip_address=request.remote_addr,
            endpoint=request.endpoint,
            method=request.method
        )
        db.session.add(audit_entry)
        db.session.commit()
    except Exception as e:
        app.logger.error(f"Failed to log action: {e}")

# Dell Warranty API Integration
def sync_dell_warranty(item):
    """
    Synchronisiert Garantie-/Vertragsdaten für ein Dell-Gerät
    Nutzt die Dell TechDirect API oder die öffentliche Warranty Status API
    """
    if not item.dell_service_tag:
        return False, "Kein Dell Service Tag vorhanden"

    try:
        # Dell Warranty Status API (öffentlich zugänglich)
        # Alternative: TechDirect API mit OAuth wenn Client ID/Secret vorhanden
        dell_client_id = os.environ.get('DELL_CLIENT_ID')
        dell_client_secret = os.environ.get('DELL_CLIENT_SECRET')

        service_tag = item.dell_service_tag.strip()

        if dell_client_id and dell_client_secret:
            # OAuth2 Flow für TechDirect API
            token_url = "https://apigtwb2c.us.dell.com/auth/oauth/v2/token"
            token_data = {
                'client_id': dell_client_id,
                'client_secret': dell_client_secret,
                'grant_type': 'client_credentials'
            }

            token_resp = requests.post(token_url, data=token_data, timeout=10)
            if token_resp.status_code != 200:
                return False, f"Dell OAuth fehlgeschlagen: {token_resp.status_code}"

            access_token = token_resp.json().get('access_token')

            # Asset Entitlement API
            api_url = f"https://apigtwb2c.us.dell.com/PROD/sbil/eapi/v5/asset-entitlements"
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Accept': 'application/json'
            }
            params = {'servicetags': service_tag}

            response = requests.get(api_url, headers=headers, params=params, timeout=10)

        else:
            # Fallback: Öffentliche Warranty Status API (keine Auth erforderlich)
            api_url = f"https://apigtwb2c.us.dell.com/PROD/sbil/eapi/v5/asset-entitlements"
            params = {
                'servicetags': service_tag,
                'apikey': 'YOUR_PUBLIC_API_KEY'  # Dell stellt öffentliche Keys bereit
            }
            response = requests.get(api_url, params=params, timeout=10)

        if response.status_code != 200:
            return False, f"Dell API Fehler: {response.status_code}"

        data = response.json()

        if not data or 'entitlements' not in data or not data['entitlements']:
            return False, "Keine Garantiedaten gefunden"

        # Erste Entitlement-Info extrahieren
        entitlement = data['entitlements'][0]

        # Update Item mit Warranty-Daten
        item.dell_warranty_status = entitlement.get('serviceLevelDescription', 'Unknown')

        # End Date parsen
        end_date_str = entitlement.get('endDate')
        if end_date_str:
            try:
                # Format: "2025-12-31T23:59:59Z"
                item.dell_warranty_end_date = datetime.strptime(end_date_str[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError) as e:
                app.logger.warning(f"Failed to parse Dell warranty end date '{end_date_str}': {e}")

        item.dell_contract_type = entitlement.get('serviceLevelGroup', 'Standard')
        item.dell_last_sync = datetime.utcnow()

        db.session.commit()

        return True, "Dell Garantiedaten erfolgreich synchronisiert"

    except requests.exceptions.RequestException as e:
        app.logger.error(f"Dell API Request Error: {e}")
        return False, f"Verbindungsfehler: {str(e)}"
    except Exception as e:
        app.logger.error(f"Dell Warranty Sync Error: {e}")
        return False, f"Fehler: {str(e)}"

# Feature 22: Granulare Berechtigungsprüfung
def has_permission(user, action, resource_type, resource_id=None):
    """Prüft ob User eine spezifische Berechtigung hat"""
    if user.role == UserRole.ADMIN:
        return True  # Admins haben immer alle Berechtigungen

    # Prüfe spezifische Berechtigung
    permission = UserPermission.query.filter_by(
        user_id=user.id,
        resource_type=resource_type,
        action=action
    ).filter(
        (UserPermission.resource_id == resource_id) | (UserPermission.resource_id.is_(None))
    ).filter(
        (UserPermission.expires_at.is_(None)) | (UserPermission.expires_at > datetime.utcnow())
    ).first()

    if permission:
        return permission.granted

    # Prüfe Rollen-basierte Berechtigung
    role_permission = UserPermission.query.filter_by(
        role=user.role,
        resource_type=resource_type,
        action=action
    ).filter(
        (UserPermission.resource_id.is_(None))
    ).first()

    if role_permission:
        return role_permission.granted

    # Default-Berechtigungen basierend auf Rolle
    default_permissions = {
        'verwaltung': ['view', 'create', 'update'],
        'mitarbeiter': ['view'],
        'kunde': ['view']
    }

    return action in default_permissions.get(user.role, [])

# Feature 23: Barcode Scanner Helper
def process_barcode_scan(barcode_data):
    """Verarbeitet gescannte Barcode-Daten"""
    # Suche Item mit diesem Barcode
    item = Item.query.filter_by(barcode=barcode_data).first()
    if item:
        return {'success': True, 'item_id': item.id, 'item': item}

    # Alternativ: Suche in Seriennummer
    item = Item.query.filter_by(serial=barcode_data).first()
    if item:
        return {'success': True, 'item_id': item.id, 'item': item}

    return {'success': False, 'message': 'Kein Item mit diesem Barcode gefunden'}

# Feature 24: RFID Tag Helper
def process_rfid_scan(rfid_tag_id):
    """Verarbeitet RFID-Tag-Scans"""
    item = Item.query.filter_by(rfid_tag_id=rfid_tag_id).first()
    if item:
        # Logge den Scan
        log_action('rfid_scan', 'item', item.id, {'rfid_tag': rfid_tag_id})
        return {'success': True, 'item_id': item.id, 'item': item}

    return {'success': False, 'message': 'Kein Item mit diesem RFID-Tag gefunden'}

# Feature 2: Wartungsprüfung
def check_maintenance_due():
    """Prüft auf fällige Wartungen und erstellt Benachrichtigungen"""
    today = date.today()

    # Finde überfällige Wartungen
    overdue_maintenances = Maintenance.query.filter(
        Maintenance.scheduled_date < today,
        Maintenance.status == 'scheduled'
    ).all()

    for maintenance in overdue_maintenances:
        maintenance.status = 'overdue'

        # Benachrichtige Admins
        admins = User.query.filter_by(role = UserRole.ADMIN).all()
        for admin in admins:
            create_notification(
                admin.id,
                'maintenance_overdue',
                'Wartung überfällig',
                f'Wartung für {maintenance.item.name} ist überfällig (fällig seit {maintenance.scheduled_date})',
                f'/maintenance/{maintenance.id}'
            )

    # Finde bald fällige Wartungen (7 Tage vorher)
    upcoming_date = today + timedelta(days=7)
    upcoming_maintenances = Maintenance.query.filter(
        Maintenance.scheduled_date.between(today, upcoming_date),
        Maintenance.status == 'scheduled'
    ).all()

    for maintenance in upcoming_maintenances:
        # Benachrichtige zugewiesene Person
        if maintenance.performed_by_id:
            create_notification(
                maintenance.performed_by_id,
                'maintenance_upcoming',
                'Wartung steht bevor',
                f'Wartung für {maintenance.item.name} ist fällig am {maintenance.scheduled_date}',
                f'/maintenance/{maintenance.id}'
            )

    db.session.commit()

# Feature 3: Reservierungskonflikte prüfen
def check_reservation_conflict(item_id, start_date, end_date, exclude_reservation_id=None):
    """Prüft ob es Konflikte mit bestehenden Reservierungen gibt"""
    query = Reservation.query.filter(
        Reservation.item_id == item_id,
        Reservation.status.in_(['pending', 'confirmed']),
        or_(
            # Neue Reservierung startet während bestehender
            Reservation.start_date.between(start_date, end_date),
            # Neue Reservierung endet während bestehender
            Reservation.end_date.between(start_date, end_date),
            # Neue Reservierung umschließt bestehende
            ((Reservation.start_date >= start_date) & (Reservation.end_date <= end_date))
        )
    )

    if exclude_reservation_id:
        query = query.filter(Reservation.id != exclude_reservation_id)

    conflicts = query.all()
    return len(conflicts) > 0, conflicts

def fetch_dell_license(serial):
    # Platzhalter für Dell-API
    return {'status': 'Aktiv', 'expiry': date(2025,12,31)}

def send_contract_email(loan):
    borrower = loan.borrower
    item     = loan.item

    # Wähle das Template je nach Rolle
    if borrower.role == UserRole.MITARBEITER:
        tpl = 'contract_employee.html'
    else:
        tpl = 'contract_customer.html'

    # Rendern und PDF erzeugen
    html    = render_template(tpl, loan=loan, borrower=borrower, item=item)
    pdf_buf = BytesIO()
    HTML(string=html).write_pdf(target=pdf_buf)
    pdf_buf.seek(0)

    # E-Mail zusammenstellen
    msg = Message(
        'Ihr Leihvertrag',
        sender    = app.config['MAIL_USERNAME'],
        recipients= [borrower.email]
    )
    msg.body = 'Im Anhang finden Sie Ihren Leihvertrag.'
    filename = f'Leihvertrag_{loan.id}.pdf'
    msg.attach(filename, 'application/pdf', pdf_buf.read())

    # Abschicken
    try:
        mail.send(msg)
    except Exception as e:
        app.logger.error(f'Fehler beim Senden der Vertrags-Mail an {borrower.email}: {e}')

def generate_loan_contract_pdf(loan_id):
    """
    Generate a professional PDF contract for a loan agreement.

    Args:
        loan_id: The ID of the loan

    Returns:
        BytesIO object containing the PDF data

    Raises:
        ValueError: If WeasyPrint is not available
        ValueError: If loan not found
    """
    # Check if WeasyPrint is available
    if not WEASYPRINT_AVAILABLE:
        raise ValueError('PDF generation is not available. WeasyPrint is required.')

    # Get loan or raise 404
    loan = Loan.query.get(loan_id)
    if not loan:
        raise ValueError(f'Loan with ID {loan_id} not found')

    # Prepare template data
    today = date.today()

    # Render HTML template with professional layout
    html = render_template(
        'contracts/loan_contract.html',
        loan=loan,
        today=today,
        app_version=APP_VERSION
    )

    # Generate PDF
    pdf_buffer = BytesIO()
    try:
        HTML(string=html).write_pdf(target=pdf_buffer)
        pdf_buffer.seek(0)
        return pdf_buffer
    except Exception as e:
        app.logger.error(f'Error generating PDF for loan {loan_id}: {e}')
        raise

# send_reset_email() migrated to routes/auth.py

def check_overdue():
    today = date.today()
    # Fix N+1 Query: Use joinedload to eagerly load relationships
    overdue = (Loan.query
               .join(User)
               .join(Item)
               .filter(Loan.return_date.is_(None), Loan.due_date < today)
               .all())

    for ln in overdue:
        bor = ln.borrower
        it  = ln.item
        nr  = f'{it.name} (SN: {it.serial})'

        # Mail an Entleiher
        try:
            m = Message(
                'Erinnerung: Rückgabe überfällig',
                sender=app.config['MAIL_USERNAME'],
                recipients=[bor.email]
            )
            m.body = (
                f'Sie haben {nr} bis {ln.due_date} ausgeliehen. '
                'Bitte geben Sie ihn umgehend zurück.'
            )
            mail.send(m)
            app.logger.info(f'Overdue reminder sent to {bor.username} for item {it.serial}')
        except Exception as e:
            app.logger.error(f'Failed to send overdue reminder to {bor.username}: {e}')

        # Mail an Admins (send once with all recipients)
        if app.config['ADMINS']:
            try:
                adm_msg = Message(
                    'Verzögerung bei Leihübergabe',
                    sender=app.config['MAIL_USERNAME'],
                    recipients=app.config['ADMINS']
                )
                adm_msg.body = (
                    f'{nr} von {bor.username} ist seit {ln.due_date} überfällig.'
                )
                mail.send(adm_msg)
                app.logger.info(f'Admin notification sent for overdue item {it.serial}')
            except Exception as e:
                app.logger.error(f'Failed to send admin notification for overdue item {it.serial}: {e}')

def check_for_update():
    """
    Ruft die neueste GitHub-Release über die API ab und
    vergleicht das Tag mit APP_VERSION.
    Gibt (is_newer, latest_version) zurück.
    """
    try:
        r = requests.get(
            "https://api.github.com/repos/BoondockSulfur/InventoryApp/releases/latest",
            timeout=3
        )
        r.raise_for_status()
        latest = r.json().get("tag_name", "")
        if latest and latest != APP_VERSION:
            return True, latest
    except Exception as e:
        app.logger.debug(f"Update-Check fehlgeschlagen: {e}")
    return False, APP_VERSION

# Scheduler: täglich um 09:00
scheduler.add_job(func=check_overdue, trigger='cron', hour=9, minute=0)
scheduler.start()

# ─── Routen ───────────────────────────────────────────────────────────────────

# Removed problematic after_request hook - use log_action() helper instead
# @app.after_request
# def log_audit(response):
#     if current_user.is_authenticated:
#         a = AuditLog(
#             user_id  = current_user.id,
#             endpoint = request.path,
#             method   = request.method
#         )
#         db.session.add(a)
#         db.session.commit()
#     return response

@app.context_processor
def inject_permissions_data():
    """Stellt Permissions-Liste und Modul-Status für alle Templates bereit."""
    if not current_user.is_authenticated:
        return {}
    try:
        # Start with permissions from user's direct role
        role = db.session.get(Role, current_user.role)
        perms = set([p.name for p in role.permissions] if role else [])

        # Add permissions from all groups the user belongs to
        for group in current_user.groups:
            for group_role in group.roles:
                for permission in group_role.permissions:
                    perms.add(permission.name)

        # Add module status
        modules = {
            'tickets_enabled': is_module_enabled('tickets'),
            'maintenance_enabled': is_module_enabled('maintenance'),
            'reservations_enabled': is_module_enabled('reservations'),
            'analytics_enabled': is_module_enabled('analytics')
        }

        return dict(permissions=list(perms), modules=modules)
    except Exception as e:
        # Rollback bei Fehlern (z.B. broken transaction)
        db.session.rollback()
        app.logger.warning(f"Failed to inject permissions: {e}")
        return dict(permissions=[], modules={})

def is_module_enabled(module_name):
    """Prüft ob ein Modul in den Einstellungen aktiviert ist"""
    setting = db.session.get(Setting, f'ENABLE_{module_name.upper()}')
    if setting:
        return setting.value.lower() in ['true', '1', 'yes', 'on']
    return True  # Standardmäßig aktiviert, wenn keine Einstellung vorhanden

def requires_permission(perm):
    """
    Decorator zur Überprüfung von Berechtigungen.
    Unterstützt sowohl einzelne Permissions als auch Listen/Tupel von Permissions.
    Bei mehreren Permissions werden diese mit OR verknüpft (eine muss erfüllt sein).
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            # Prüfe zuerst ob der Benutzer eingeloggt ist
            if not current_user.is_authenticated:
                flash('Bitte melden Sie sich an, um diese Aktion durchzuführen.', 'warning')
                abort(HTTPStatus.UNAUTHORIZED)  # Unauthorized

            try:
                # Erlaube mehrere Permissions (OR-Verknüpfung)
                permissions = perm if isinstance(perm, (list, tuple)) else [perm]

                # Prüfe mit der neuen User-Methode
                if current_user.has_any_permission(*permissions):
                    return f(*args, **kwargs)

                # Keine passende Berechtigung gefunden
                app.logger.warning(
                    f"User {current_user.username} (role: {current_user.role}) "
                    f"tried to access {f.__name__} without permission {permissions}"
                )
                flash('Sie haben keine Berechtigung für diese Aktion.', FlashCategory.DANGER)
                abort(HTTPStatus.FORBIDDEN)

            except Exception as e:
                app.logger.error(f"Permission check failed for {f.__name__}: {e}")
                # Admin kann immer zugreifen, sonst 403
                if current_user.is_authenticated and current_user.is_admin:
                    return f(*args, **kwargs)
                flash('Ein Fehler ist bei der Berechtigungsprüfung aufgetreten.', FlashCategory.DANGER)
                abort(HTTPStatus.FORBIDDEN)
        return wrapped
    return decorator

def requires_any_permission(*perms):
    """
    Decorator der prüft ob der Benutzer mindestens eine der angegebenen Berechtigungen hat.
    Beispiel: @requires_any_permission('manage_items', 'view_items')
    """
    return requires_permission(perms)

def requires_all_permissions(*perms):
    """
    Decorator der prüft ob der Benutzer alle angegebenen Berechtigungen hat.
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Bitte melden Sie sich an, um diese Aktion durchzuführen.', 'warning')
                abort(HTTPStatus.UNAUTHORIZED)

            try:
                if current_user.has_all_permissions(*perms):
                    return f(*args, **kwargs)

                app.logger.warning(
                    f"User {current_user.username} tried to access {f.__name__} "
                    f"but missing some of {perms}"
                )
                flash('Sie haben nicht alle erforderlichen Berechtigungen für diese Aktion.', FlashCategory.DANGER)
                abort(HTTPStatus.FORBIDDEN)

            except Exception as e:
                app.logger.error(f"Permission check failed: {e}")
                if current_user.is_authenticated and current_user.is_admin:
                    return f(*args, **kwargs)
                abort(HTTPStatus.FORBIDDEN)
        return wrapped
    return decorator

def admin_required(f):
    """Decorator der nur Admins Zugriff gewährt."""
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Bitte melden Sie sich an.', FlashCategory.WARNING)
            abort(HTTPStatus.UNAUTHORIZED)

        if not current_user.is_admin:
            flash('Nur Administratoren haben Zugriff auf diese Funktion.', FlashCategory.DANGER)
            abort(HTTPStatus.FORBIDDEN)

        return f(*args, **kwargs)
    return wrapped

@app.context_processor
def inject_permissions():
    """
    Stellt Permission-Helper-Funktionen für alle Templates bereit.
    """
    def has_permission(perm):
        """Template-Helper: Prüft ob aktueller User eine Permission hat"""
        if not current_user.is_authenticated:
            return False
        return current_user.has_permission(perm)

    def has_any_permission(*perms):
        """Template-Helper: Prüft ob aktueller User mindestens eine Permission hat"""
        if not current_user.is_authenticated:
            return False
        return current_user.has_any_permission(*perms)

    def has_all_permissions(*perms):
        """Template-Helper: Prüft ob aktueller User alle Permissions hat"""
        if not current_user.is_authenticated:
            return False
        return current_user.has_all_permissions(*perms)

    def is_admin():
        """Template-Helper: Prüft ob aktueller User Admin ist"""
        return current_user.is_authenticated and current_user.is_admin

    def is_manager():
        """Template-Helper: Prüft ob aktueller User Manager ist"""
        return current_user.is_authenticated and current_user.is_manager

    def user_permissions():
        """Template-Helper: Gibt alle Permissions des aktuellen Users zurück"""
        if not current_user.is_authenticated:
            return set()
        return current_user.get_all_permissions()

    return {
        'has_permission': has_permission,
        'has_any_permission': has_any_permission,
        'has_all_permissions': has_all_permissions,
        'is_admin': is_admin,
        'is_manager': is_manager,
        'user_permissions': user_permissions
    }

@app.context_processor
def inject_update_info():
    """
    Stellt Update-Informationen für alle Templates bereit.
    Check wird nur für eingeloggte Admins durchgeführt.
    """
    from flask import session

    if not current_user.is_authenticated or current_user.role != 'admin':
        return {}

    is_newer, latest = check_for_update()

    # Prüfen ob User diese Version bereits dismissed hat
    dismissed_version = session.get('dismissed_version', '')
    show_update = is_newer and (dismissed_version != latest)

    return {
        "update_available": show_update,
        "latest_version": latest,
        "current_version": APP_VERSION
    }

# ==================== Enhanced Ticket System Initialization ====================
# Initialize enhanced ticket system routes only (models exist in database)
try:
    from ticket_system import init_ticket_routes, ticket_bp

    # Register ticket system blueprint with routes
    init_ticket_routes(app, db, requires_permission)

    # SECURITY NOTE: CSRF protection is now ENABLED for ticket routes
    # AJAX requests must include CSRF token in X-CSRFToken header
    # csrf.exempt(ticket_bp) # REMOVED for security

    app.logger.info("✅ Ticket system routes registered")
except Exception as e:
    app.logger.error(f"Ticket System initialization failed: {e}")
    import traceback
    app.logger.error(traceback.format_exc())

# ==================== Security Headers ====================
@app.after_request
def set_security_headers(response):
    """Set security headers for all responses"""
    # Prevent MIME-sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'

    # Prevent clickjacking
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'

    # Enable browser XSS protection
    response.headers['X-XSS-Protection'] = '1; mode=block'

    # Only set HSTS for HTTPS connections
    if request.is_secure:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    return response

# ==================== Routes ====================

@app.route('/api/dismiss_update', methods=['POST'])
@login_required
def dismiss_update():
    """Speichert, dass User Update-Benachrichtigung nicht mehr sehen möchte"""
    from flask import session
    data = request.get_json() or {}
    session['update_dismissed'] = True
    session['dismissed_version'] = data.get('version', '')
    return jsonify({'success': True})

@app.route('/generate_serial', methods=['GET'])
@login_required
def generate_serial():
    """Generate a new serial number without automatically printing.

    Returns the serial number in JSON format. User can then use the label
    designer to create and print a custom label.
    """
    new_serial = uuid.uuid4().hex[:8].upper()
    # Return serial number without printing - user will design label manually
    return jsonify({
        'serial': new_serial,
        'label_designer_url': url_for('labels.generator', serial=new_serial)
    })
# Auth route 'Login route' migrated to routes/auth.py
# Auth route 'OAuth routes' migrated to routes/auth.py

# Auth routes 'SAML routes + helpers' migrated to routes/auth.py

# Auth routes 'Password reset routes' migrated to routes/auth.py



# ============================================================================
# LOAN ROUTES - Restored from Git
# ============================================================================

@app.route('/item/<int:item_id>/loan', methods=['GET','POST'])
@login_required
@requires_permission('manage_items')
def loan_item(item_id):
    # Gegenstand laden oder 404
    itm = Item.query.get_or_404(item_id)

    # Mögliche Entleiher (Mitarbeiter und Kunden)
    borrowers = User.query.filter(
        User.role.in_(['mitarbeiter', 'kunde'])
    ).all()

    # Formular initialisieren
    form = DummyForm()
    if form.validate_on_submit():
        # POST-Logik: ausgewählten User finden
        bor = User.query.filter_by(username=form.borrower_name.data.strip()).first()
        if not bor:
            flash('Entleiher nicht gefunden', 'danger')
            return redirect(url_for('loan_item', item_id=item_id))

        if form.due_date.data < date.today():
            flash('Rückgabedatum in der Vergangenheit', 'warning')
            return redirect(url_for('loan_item', item_id=item_id))

        # Leihvorgang anlegen
        ln = Loan(item_id=itm.id, borrower_id=bor.id, due_date=form.due_date.data)
        itm.is_borrowed = True
        db.session.add(ln)
        db.session.commit()
        broadcast_socketio('loan_created', {'id': ln.id, 'item_id': itm.id, 'borrower_id': bor.id})

        flash('Leihvorgang erstellt', 'success')
        send_contract_email(ln)
        return redirect(url_for('item_detail', item_id=item_id))

    # GET-Request oder fehlerhafte POST-Validierung: Formular anzeigen
    return render_template(
        'loan_form.html',
        item=itm,
        borrowers=borrowers,
        form=form
    )

@app.route('/item/<int:item_id>/return', methods=['POST'])
@login_required
def return_item(item_id):
    form = ReturnForm()
    if not form.validate_on_submit():
        flash('Ungültige Anfrage','danger')
        return redirect(url_for('item_detail', item_id=item_id))
    if current_user.role not in ('admin','verwaltung'):
        flash('Zugriff verweigert','danger')
        return redirect(url_for('items'))
    itm = Item.query.get_or_404(item_id)
    ln  = Loan.query.filter_by(item_id=itm.id, return_date=None).first()
    if not ln:
        flash('Kein aktiver Leihvorgang','warning')
        return redirect(url_for('item_detail', item_id=item_id))
    ln.return_date   = date.today()
    itm.is_borrowed = False
    db.session.commit()
    broadcast_socketio('loan_returned', {'id': ln.id, 'item_id': itm.id})
    flash('Rückgabe verbucht','success')
    return redirect(url_for('item_detail', item_id=item_id))

@app.route('/item/<int:item_id>/print_label')
@login_required
def print_label(item_id):
    itm = Item.query.get_or_404(item_id)
    zpl = f'^XA^FO50,50^BCN,100,Y,N,N^FD{itm.serial}^FS^XZ'
    try:
        with socket.socket() as s:
            s.connect((app.config['PRINTER_IP'], app.config['PRINTER_PORT']))
            s.send(zpl.encode('utf-8'))
        flash('Label an Drucker gesendet','success')
    except Exception:
        flash('Druck fehlgeschlagen','danger')
    return redirect(url_for('item_detail', item_id=item_id))

@app.route('/item/<int:item_id>/defect', methods=['POST'])
@login_required
def mark_defective(item_id):
    if current_user.role not in ('admin','verwaltung'):
        abort(403)
    itm = Item.query.get_or_404(item_id)
    itm.defective = True
    itm.is_borrowed = True   # gleichzeitig nicht verfügbar
    db.session.commit()
    flash('Gegenstand als defekt markiert','warning')
    return redirect(url_for('item_detail', item_id=item_id))

@app.route('/item/<int:item_id>/repair', methods=['POST'])
@login_required
def mark_repaired(item_id):
    if current_user.role not in ('admin','verwaltung'):
        abort(403)
    itm = Item.query.get_or_404(item_id)
    itm.defective = False
    itm.is_borrowed = False  # wieder verfügbar
    db.session.commit()
    flash('Gegenstand als repariert markiert','success')
    return redirect(url_for('item_detail', item_id=item_id))

@app.route('/loans/active')
@login_required
def active_loans():
    loans = Loan.query.filter_by(return_date=None).all()
    today = date.today()
    return render_template('active_loans.html', loans=loans, today=today)

@app.route('/users')
@login_required
@requires_permission('manage_users')
def users():
    if current_user.role not in ('admin', 'verwaltung'):
        flash('Zugriff verweigert', 'danger')
        return redirect(url_for('dashboard'))
    q = request.args.get('q', '').strip()
    query = User.query
    if q:
        query = query.filter(
            or_(
                User.username.ilike(f'%{q}%'),
                User.email.ilike(f'%{q}%'),
                User.role.ilike(f'%{q}%')
            )
        )
    users_list = query.order_by(User.id).all()
    return render_template('users.html', users=users_list, search=q)

@app.route('/user/add', methods=['GET','POST'])
@login_required
@requires_permission('manage_users')
def add_user():
    if current_user.role not in ('admin','verwaltung'):
        flash('Zugriff verweigert', 'danger')
        return redirect(url_for('dashboard'))

    form = RegistrationForm()
    if form.validate_on_submit():
        # Neuen User anlegen
        u = User(
            username     = form.username.data,
            email        = form.email.data,
            role         = form.role.data,
            group_number = form.group_number.data if form.role.data == 'kunde' else None
        )
        u.set_password(form.password.data)
        db.session.add(u)
        db.session.commit()

        flash('Benutzerkonto erstellt', 'success')
        return redirect(url_for('users'))

    return render_template('add_user.html', form=form)

@app.route('/contract/<int:loan_id>')
@login_required
def contract(loan_id):
    ln  = Loan.query.get_or_404(loan_id)
    bor = ln.borrower
    it  = ln.item

    # Wähle Template anhand der User-Rolle
    if bor.role == 'mitarbeiter':
        tpl = 'contract_employee.html'
    else:
        tpl = 'contract_customer.html'

    # Check if PDF generation is available
    if not WEASYPRINT_AVAILABLE:
        flash('PDF-Generierung ist in dieser Installation nicht verfügbar. Bitte nutzen Sie das Docker-Setup für alle Features.', 'warning')
        # Return HTML version instead
        return render_template(tpl, loan=ln, borrower=bor, item=it)

    # HTML rendern und PDF erzeugen
    html    = render_template(tpl, loan=ln, borrower=bor, item=it)
    pdf_buf = BytesIO()
    HTML(string=html).write_pdf(target=pdf_buf)
    pdf_buf.seek(0)

    # PDF ausliefern
    return send_file(
        pdf_buf,
        download_name=f'Leihvertrag_{ln.id}.pdf',
        as_attachment=True
    )

@app.route('/meine_leihgaben')
@login_required
@requires_permission('view_own_loans')
def meine_leihgaben():
    if current_user.role != 'kunde':
        flash('Zugriff verweigert','danger')
        return redirect(url_for('dashboard'))
    # Borrower-Datensatz finden (nach E-Mail)
    bor = User.query.filter_by(email=current_user.email).first()
    if not bor:
        flash('Kein Entleiherprofil gefunden','warning')
        return redirect(url_for('dashboard'))
    # nur die eigenen Leihvorgänge
    loans = Loan.query.filter_by(borrower_id=bor.id).all()
    return render_template('meine_leihgaben.html', loans=loans)

@app.route('/tickets')
@login_required
@requires_permission('view_own_tickets')
def my_tickets():
    if not is_module_enabled('tickets'):
        flash('Das Ticketsystem ist derzeit deaktiviert', 'warning')
        return redirect(url_for('dashboard'))
    # alle Tickets des eingeloggten Users
    tickets = Ticket.query.filter_by(user_id=current_user.id).order_by(Ticket.created_at.desc()).all()
    return render_template('tickets.html', tickets=tickets)

# ============================================================================
# ADMIN SETTINGS - Restored from Git
# ============================================================================
@app.route('/admin/settings', methods=['GET','POST'])
@login_required
@requires_permission('manage_settings')
def admin_settings():
    form = SettingsForm()

    # Alle Settings aus DB laden
    settings_dict = {s.key: s.value for s in Setting.query.all()}

    # Initialbefüllung aus DB
    if request.method == 'GET':
        form.mail_server.data   = settings_dict.get('MAIL_SERVER')
        form.mail_port.data     = int(settings_dict.get('MAIL_PORT',0))
        form.mail_use_tls.data  = settings_dict.get('MAIL_USE_TLS','false').lower() in ['true','1','yes']
        form.mail_username.data = settings_dict.get('MAIL_USERNAME')
        form.admins.data        = settings_dict.get('ADMINS')
        form.printer_ip.data    = settings_dict.get('PRINTER_IP')
        form.printer_port.data  = int(settings_dict.get('PRINTER_PORT',0))
        form.dell_client_id.data     = settings_dict.get('DELL_CLIENT_ID')

    if request.method == 'POST':
        # Bestimme welcher Tab gespeichert wird
        tab = request.form.get('tab', 'email')

        # Sammle alle Form-Daten (außer CSRF-Token und tab)
        updates = {}

        # Define checkbox fields first so we can handle them specially
        checkbox_fields = [
            'mail_use_tls', 'ldap_enabled', 'enable_qr', 'enable_barcode', 'enable_rfid',
            'auto_maintenance_reminders', 'enable_budget_tracking', 'require_2fa',
            'require_strong_passwords', 'auto_backup_enabled', 'notify_new_items',
            'notify_loans', 'notify_maintenance', 'notify_tickets',
            'enable_tickets', 'enable_maintenance', 'enable_reservations', 'enable_analytics',
            'oauth_auto_register', 'saml_auto_register', 'enable_multi_company'
        ]

        for key, value in request.form.items():
            if key not in ['csrf_token', 'tab', 'submit']:
                # Checkboxen normalisieren zu 'true'/'false'
                if key in checkbox_fields:
                    # HTML checkboxes send 'on' when checked, normalize to 'true'
                    updates[key.upper()] = 'true' if value in ['on', 'true', '1', 'yes'] else 'false'
                else:
                    updates[key.upper()] = value

        # Checkbox-Werte für nicht-gecheckte Checkboxen setzen (unchecked boxes don't appear in request.form)
        for checkbox in checkbox_fields:
            if checkbox not in request.form and checkbox.upper() not in updates:
                updates[checkbox.upper()] = 'false'

        # WTForms-basierte Updates (für E-Mail und Import/Export Tab)
        if tab in ['email', 'import'] and form.validate_on_submit():
            wtf_updates = {
                'MAIL_SERVER':  form.mail_server.data,
                'MAIL_PORT':    str(form.mail_port.data),
                'MAIL_USE_TLS': 'true' if form.mail_use_tls.data else 'false',
                'MAIL_USERNAME':form.mail_username.data,
                'MAIL_PASSWORD':form.mail_password.data if form.mail_password.data else settings_dict.get('MAIL_PASSWORD', ''),
                'ADMINS':       form.admins.data,
                'PRINTER_IP':   form.printer_ip.data,
                'PRINTER_PORT': str(form.printer_port.data),
                'DELL_CLIENT_ID':    form.dell_client_id.data,
                'DELL_CLIENT_SECRET':form.dell_client_secret.data if form.dell_client_secret.data else settings_dict.get('DELL_CLIENT_SECRET', '')
            }
            updates.update(wtf_updates)

        # Settings in DB speichern/aktualisieren
        for key, value in updates.items():
            key_upper = key.upper()
            setting = db.session.get(Setting, key_upper)
            if setting:
                setting.value = str(value)
            else:
                # Neues Setting anlegen
                db.session.add(Setting(key=key_upper, value=str(value)))

        db.session.commit()
        flash('Einstellungen gespeichert', 'success')

        # Zurück zum gleichen Tab mit Hash
        return redirect(url_for('admin_settings') + f'#{tab}')

    # Logs laden
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(50).all()
    log_path   = app.config.get('ERROR_LOG_PATH', os.path.join(app.instance_path,'error.log'))
    try:
        with open(log_path, encoding='utf-8') as f:
            errors = f.readlines()[-50:]
    except FileNotFoundError:
        errors = []

    # Check which OAuth providers are configured
    oauth_providers_config = {
        'google': bool(os.environ.get('GOOGLE_CLIENT_ID') and os.environ.get('GOOGLE_CLIENT_SECRET')),
        'microsoft': bool(os.environ.get('MICROSOFT_CLIENT_ID') and os.environ.get('MICROSOFT_CLIENT_SECRET')),
        'github': bool(os.environ.get('GITHUB_CLIENT_ID') and os.environ.get('GITHUB_CLIENT_SECRET'))
    }

    return render_template(
      'admin_settings.html',
      form       = form,
      settings   = settings_dict,
      audits     = audit_logs,
      errors     = errors,
      oauth_providers_config = oauth_providers_config
    )

@app.route('/dashboard')
@login_required
def dashboard():
    total    = Item.query.count()
    borrowed = Item.query.filter_by(is_borrowed=True).count()
    # Fix N+1 Query: Eager load item and borrower for overdue loans
    overdue  = Loan.query.options(
        joinedload(Loan.item),
        joinedload(Loan.borrower)
    ).filter(Loan.return_date.is_(None), Loan.due_date<date.today()).all()
    return render_template(
        'dashboard.html',
        total_items=total,
        borrowed_items=borrowed,
        overdue_loans=overdue
    )

# ============================================================================
# NEW FEATURES: Advanced Search Route
# ============================================================================

@app.route('/advanced-search')
@login_required
def elasticsearch_search():
    """Advanced search page with Elasticsearch"""
    return render_template('advanced_search.html')


@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    """Get all unique categories for filter dropdown"""
    try:
        categories = db.session.query(Item.category).distinct().filter(Item.category.isnot(None)).all()
        category_list = [cat[0] for cat in categories if cat[0]]
        return jsonify({'success': True, 'categories': category_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/items')
@login_required
@requires_permission('manage_items')
def items():
    # Get filter parameters with validation
    q = request.args.get('q', '').strip()

    # Limit search query length to prevent abuse
    if len(q) > 200:
        flash('Suchbegriff zu lang (max. 200 Zeichen)', FlashCategory.WARNING)
        q = q[:200]

    # Validate category_filter is an integer if provided
    category_filter = request.args.get('category', '')
    if category_filter:
        try:
            category_filter = int(category_filter)
        except (ValueError, TypeError):
            flash('Ungültige Kategorie-ID', FlashCategory.WARNING)
            category_filter = ''

    # Validate location_filter is an integer if provided
    location_filter = request.args.get('location', '')
    if location_filter:
        try:
            location_filter = int(location_filter)
        except (ValueError, TypeError):
            flash('Ungültige Standort-ID', FlashCategory.WARNING)
            location_filter = ''

    # Validate status_filter is from allowed values
    status_filter = request.args.get('status', '')
    allowed_statuses = ['available', 'borrowed', 'defective', '']
    if status_filter not in allowed_statuses:
        flash('Ungültiger Status-Filter', FlashCategory.WARNING)
        status_filter = ''

    query = Item.query

    # Role-based access control
    if current_user.role not in ('admin', 'verwaltung'):
        flash('Zugriff verweigert', FlashCategory.DANGER)
        # Fix N+1 Query: Use subquery instead of loading all loans
        my_item_ids = db.session.query(Loan.item_id).filter_by(borrower_id=current_user.id).subquery()
        query = query.filter(Item.id.in_(my_item_ids))

    # Apply search filter
    if q:
        query = query.filter(
            db.or_(
                Item.name.ilike(f'%{q}%'),
                Item.serial.ilike(f'%{q}%'),
                Item.description.ilike(f'%{q}%')
            )
        )

    # Apply category filter
    if category_filter:
        query = query.filter(Item.category_id == int(category_filter))

    # Apply location filter
    if location_filter:
        query = query.filter(Item.location_id == int(location_filter))

    # Apply status filter
    if status_filter == 'available':
        query = query.filter(Item.is_borrowed == False, Item.defective == False)
    elif status_filter == 'borrowed':
        query = query.filter(Item.is_borrowed == True)
    elif status_filter == 'defective':
        query = query.filter(Item.defective == True)

    # Get all categories and locations for filter dropdowns
    all_categories = Category.query.order_by(Category.name).all()
    all_locations = Location.query.order_by(Location.name).all()

    # Add pagination to avoid loading all items at once
    # Fix N+1 Query: Eager load related data
    page = request.args.get('page', 1, type=int)
    per_page = 50
    query = query.options(
        joinedload(Item.category_rel),
        joinedload(Item.location_rel),
        selectinload(Item.loans)
    )
    pagination = query.order_by(Item.is_borrowed.asc(), Item.name.asc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template(
        'items.html',
        items=pagination.items,
        pagination=pagination,
        search=q,
        all_categories=all_categories,
        all_locations=all_locations
    )

@app.route('/items/export/csv')
@login_required
@requires_permission('manage_items')
def export_items_csv():
    """Exportiert alle Items als CSV"""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(['ID', 'Name', 'Seriennummer', 'Standort', 'Kategorie', 'Status', 'Notizen', 'Erstellt am'])

    # Daten
    items = Item.query.order_by(Item.name).all()
    for item in items:
        status = 'Defekt' if item.defective else ('Verliehen' if item.is_borrowed else 'Verfügbar')
        writer.writerow([
            item.id,
            item.name,
            item.serial,
            item.location or '',
            item.category,
            status,
            item.note or '',
            item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else ''
        ])

    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8-sig')),  # UTF-8 mit BOM für Excel
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'inventory_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

@app.route('/items/export/excel')
@login_required
@requires_permission('manage_items')
def export_items_excel():
    """Exportiert alle Items als Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Header mit Styling
    headers = ['ID', 'Name', 'Seriennummer', 'Standort', 'Kategorie', 'Status', 'Notizen', 'Erstellt am']
    ws.append(headers)

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Daten
    items = Item.query.order_by(Item.name).all()
    for item in items:
        status = 'Defekt' if item.defective else ('Verliehen' if item.is_borrowed else 'Verfügbar')
        ws.append([
            item.id,
            item.name,
            item.serial,
            item.location or '',
            item.category,
            status,
            item.note or '',
            item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else ''
        ])

    # Auto-width für Spalten
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, AttributeError) as e:
                app.logger.debug(f"Excel column width calculation error for cell: {e}")
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Speichern in BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventory_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@app.route('/items/bulk-import')
@login_required
@requires_permission('manage_items')
def bulk_import_page():
    """Professioneller Bulk Import mit Column Mapping"""
    return render_template('items/bulk_import_improved.html')

@app.route('/api/bulk-import/template/csv')
@login_required
@requires_permission('manage_items')
def bulk_import_template_csv():
    """Download CSV template for bulk import"""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(['name', 'serial', 'category', 'location', 'note', 'barcode', 'purchase_price', 'purchase_date'])
    writer.writerow(['Beispiel Laptop', 'SN-001', 'IT-Equipment', 'Büro 1', 'Neuanschaffung', 'BC-001', '999.99', '2024-01-15'])

    output.seek(0)
    return send_file(
        BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name='inventory_import_template.csv'
    )

@app.route('/api/bulk-import/template/xlsx')
@login_required
@requires_permission('manage_items')
def bulk_import_template_xlsx():
    """Download Excel template for bulk import"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"

    headers = ['name', 'serial', 'category', 'location', 'note', 'barcode', 'purchase_price', 'purchase_date']
    ws.append(headers)
    ws.append(['Beispiel Laptop', 'SN-001', 'IT-Equipment', 'Büro 1', 'Neuanschaffung', 'BC-001', '999.99', '2024-01-15'])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='inventory_import_template.xlsx'
    )

@app.route('/api/bulk-import/items', methods=['POST'])
@login_required
@requires_permission('manage_items')
@limiter.limit("10 per hour")  # Rate limit for bulk operations
def bulk_import_items():
    """Process bulk import of items"""
    try:
        data = request.get_json()
        items_data = data.get('items', [])

        success_count = 0
        error_count = 0
        errors = []

        for item_data in items_data:
            try:
                # Check for required fields
                if not item_data.get('name') or not item_data.get('serial'):
                    errors.append(f"Missing required fields for item: {item_data}")
                    error_count += 1
                    continue

                # Check for duplicate serial
                existing = Item.query.filter_by(serial=item_data['serial']).first()
                if existing:
                    errors.append(f"Duplicate serial: {item_data['serial']}")
                    error_count += 1
                    continue

                # Create item
                item = Item(
                    name=item_data['name'],
                    serial=item_data['serial'],
                    category=item_data.get('category', 'other'),
                    location=item_data.get('location'),
                    note=item_data.get('note'),
                    barcode=item_data.get('barcode'),
                    purchase_price=float(item_data['purchase_price']) if item_data.get('purchase_price') else None,
                    purchase_date=datetime.strptime(item_data['purchase_date'], '%Y-%m-%d').date() if item_data.get('purchase_date') else None
                )

                db.session.add(item)
                success_count += 1

            except Exception as e:
                errors.append(f"Error importing {item_data.get('serial', 'unknown')}: {str(e)}")
                error_count += 1

        db.session.commit()
        log_action('bulk_import', 'item', None, {'count': success_count})

        return jsonify({
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/item/add', methods=['GET','POST'])
@login_required
@requires_permission('manage_items')
def add_item():
    if current_user.role not in ('admin','verwaltung'):
        flash('Zugriff verweigert', FlashCategory.DANGER)
        return redirect(url_for('items'))

    form = ItemForm()

    # Populate dropdown choices
    categories = Category.query.order_by(Category.name).all()
    locations = Location.query.order_by(Location.name).all()
    items = Item.query.order_by(Item.name).all()

    form.category_id.choices = [(None, '-- Keine Kategorie --')] + [(c.id, c.name) for c in categories]
    form.location_id.choices = [(None, '-- Kein Standort --')] + [(l.id, l.name) for l in locations]
    form.parent_item_id.choices = [(None, '-- Kein übergeordneter Gegenstand --')] + [(i.id, f"{i.name} ({i.serial})") for i in items]

    # Populate budget choices
    current_year = date.today().year
    budgets = BudgetEntry.query.filter_by(year=current_year).order_by(BudgetEntry.name).all()
    form.budget_id.choices = [(None, '-- Kein Budget --')] + [(b.id, f"{b.name} ({b.get_remaining_budget():.2f}€ verfügbar)") for b in budgets]

    if form.validate_on_submit():
        try:
            # 1) Neuen Gegenstand anlegen und speichern
            itm = Item(
                name=form.name.data,
                serial=form.serial.data,
                note=form.note.data,
                category_id=form.category_id.data,
                location_id=form.location_id.data,
                parent_item_id=form.parent_item_id.data,
                purchase_price=float(form.purchase_price.data) if form.purchase_price.data else None,
                purchase_date=form.purchase_date.data
            )

            # Set current_value to purchase_price initially
            if itm.purchase_price:
                itm.current_value = itm.purchase_price

            # Set old fields for backward compatibility
            if form.category_id.data:
                cat = Category.query.filter_by(id=form.category_id.data).first()
                if cat:
                    itm.category = cat.name.lower()

            if form.location_id.data:
                loc = Location.query.filter_by(id=form.location_id.data).first()
                if loc:
                    itm.location = loc.name

            db.session.add(itm)
            db.session.flush()  # Get item ID for budget transaction

            # Handle budget deduction if requested
            if form.deduct_from_budget.data and form.budget_id.data and form.purchase_price.data:
                budget = BudgetEntry.query.get(form.budget_id.data)
                if budget:
                    transaction = BudgetTransaction(
                        budget_id=budget.id,
                        amount=float(form.purchase_price.data),
                        transaction_type='purchase',
                        transaction_date=form.purchase_date.data or date.today(),
                        description=f"Kauf: {itm.name} (S/N: {itm.serial})",
                        reference_type='item',
                        reference_id=itm.id,
                        created_by=current_user.id
                    )
                    db.session.add(transaction)
                    db.session.flush()  # Ensure transaction is in session for calculation
                    budget.spent_amount = budget.calculate_spent_amount()
                    flash(f'Betrag von {form.purchase_price.data}€ von Budget "{budget.name}" abgezogen', FlashCategory.INFO)

            db.session.commit()
            broadcast_socketio('item_created', {'id': itm.id, 'name': itm.name})
            flash('Gegenstand hinzugefügt', FlashCategory.SUCCESS)
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to add item: {e}")
            flash('Fehler beim Hinzufügen des Gegenstands', FlashCategory.DANGER)
            return render_template('add_item.html', form=form)

        # 2) Nur für Laptops: Service-Tag in CSV + CLI ausführen
        if itm.category == 'laptop':
            # Arbeitsverzeichnis unter instance/warranty
            csv_dir = app.config.get('WARRANTY_CSV_DIR') \
                      or os.path.join(app.instance_path, 'warranty')
            os.makedirs(csv_dir, exist_ok=True)

            input_csv  = os.path.join(csv_dir, 'Dell-Support.csv')
            output_csv = os.path.join(csv_dir, 'Dell-Support-Ausgabe.csv')

            # Seriennummer anhängen
            with open(input_csv, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow([itm.serial])

            # CLI aufrufen
            cli_cmd = app.config.get('WARRANTY_CLI_CMD', 'DellWarranty-CLI.exe')
            try:
                subprocess.run(
                    [cli_cmd, f'/I={input_csv}', f'/E={output_csv}'],
                    check=True
                )
            except subprocess.CalledProcessError as e:
                app.logger.error(f'Warranty-CLI-Fehler (exit {e.returncode})')
            except FileNotFoundError:
                app.logger.error(f'Warranty-CLI nicht gefunden: {cli_cmd}')
            except Exception as e:
                app.logger.error(f'Unerwarteter Fehler bei Warranty-CLI: {e}')

        return redirect(url_for('items'))

    # Bei GET oder ungültigem Formular zurück zum Formular
    return render_template('add_item.html', form=form)

@app.route('/item/<int:item_id>/edit', methods=['GET','POST'])
@login_required
@requires_permission('manage_items')
def edit_item(item_id):
    itm = Item.query.get_or_404(item_id)
    if current_user.role != 'admin':
        abort(HTTPStatus.FORBIDDEN)

    form = ItemForm(obj=itm)

    # Populate dropdown choices
    categories = Category.query.order_by(Category.name).all()
    locations = Location.query.order_by(Location.name).all()
    # Exclude current item from parent options to prevent circular references
    items = Item.query.filter(Item.id != item_id).order_by(Item.name).all()

    form.category_id.choices = [(None, '-- Keine Kategorie --')] + [(c.id, c.name) for c in categories]
    form.location_id.choices = [(None, '-- Kein Standort --')] + [(l.id, l.name) for l in locations]
    form.parent_item_id.choices = [(None, '-- Kein übergeordneter Gegenstand --')] + [(i.id, f"{i.name} ({i.serial})") for i in items]

    if form.validate_on_submit():
        try:
            # Update item fields
            itm.name = form.name.data
            itm.serial = form.serial.data
            itm.note = form.note.data
            itm.category_id = form.category_id.data
            itm.location_id = form.location_id.data
            itm.parent_item_id = form.parent_item_id.data

            # Update old fields for backward compatibility
            if form.category_id.data:
                cat = Category.query.filter_by(id=form.category_id.data).first()
                if cat:
                    itm.category = cat.name.lower()
            else:
                itm.category = None

            if form.location_id.data:
                loc = Location.query.filter_by(id=form.location_id.data).first()
                if loc:
                    itm.location = loc.name
            else:
                itm.location = None

            db.session.commit()
            broadcast_socketio('item_updated', {'id': itm.id, 'name': itm.name})
            flash('Gegenstand aktualisiert', FlashCategory.SUCCESS)
            return redirect(url_for('item_detail', item_id=itm.id))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to update item {item_id}: {e}")
            flash('Fehler beim Aktualisieren des Gegenstands', FlashCategory.DANGER)

    # Pre-populate form fields on GET request
    if request.method == 'GET':
        form.category_id.data = itm.category_id
        form.location_id.data = itm.location_id
        form.parent_item_id.data = itm.parent_item_id

    return render_template('edit_item.html', form=form, item=itm)

@app.route('/item/<int:item_id>/delete', methods=['POST'])
@login_required
@requires_permission('manage_items')
def delete_item(item_id):
    itm = Item.query.get_or_404(item_id)
    if current_user.role != 'admin':
        abort(HTTPStatus.FORBIDDEN)
    try:
        item_name = itm.name
        db.session.delete(itm)
        db.session.commit()
        broadcast_socketio('item_deleted', item_id)
        flash('Gegenstand gelöscht', FlashCategory.WARNING)
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to delete item {item_id}: {e}")
        flash('Fehler beim Löschen des Gegenstands', FlashCategory.DANGER)
    return redirect(url_for('items'))

@app.route('/item/<int:item_id>')
@login_required
@requires_permission('manage_items')
def item_detail(item_id):
    itm         = Item.query.get_or_404(item_id)
    lic         = fetch_dell_license(itm.serial)
    act_loan    = Loan.query.filter_by(item_id=itm.id, return_date=None).first()
    return_form = ReturnForm()
    today       = date.today()
    warranty = None
    if itm.category == 'laptop':
        csv_dir = app.config.get('WARRANTY_CSV_DIR') \
                  or os.path.join(app.instance_path, 'warranty'
        )
        output_csv = os.path.join(csv_dir, 'Dell-Support-Ausgabe.csv')
        try:
            with open(output_csv, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if row.get('ServiceTag') == itm.serial:
                        warranty = row
                        break
        except FileNotFoundError:
            warranty = None
    return render_template(
        'item_detail.html',
        item=itm, loan=act_loan, license=lic,
        return_form=return_form, today=today
    )

# ============================================================================
# ADMIN ROUTES - MIGRATED TO BLUEPRINT
# ============================================================================
# All admin routes have been moved to routes/admin.py
# The admin blueprint is initialized at the end of this file (line ~6100)
#
# Migrated routes (20 routes, ~1155 lines):
# - /admin/categories (GET, POST, PUT, DELETE)
# - /admin/locations (GET, POST, PUT, DELETE)
# - /admin/tickets (GET, POST)
# - /admin/roles (GET, POST, PUT, DELETE)
# - /admin/system-logs (GET)
# - /admin/settings (GET, POST)
# - /admin/setup (GET)
# - /admin/check-updates (GET)
# - /admin/apply-update (POST)
# - /admin/restart-system (POST)
# ============================================================================

# ============================================================================
# ─── ML API Routes ─────────────────────────────────────────────────────────

@app.route('/api/ml/stats', methods=['GET'])
@login_required
@requires_permission('manage_settings')
def ml_stats():
    """Get ML statistics and model status"""
    try:
        # Count categorized items
        categorized_items = Item.query.filter(Item.category_id.isnot(None)).count()

        # Count categories
        total_categories = Category.query.count()

        # Check if model is loaded
        model_loaded = False
        model_accuracy = None

        if hasattr(app, 'ml_service') and app.ml_service:
            model_loaded = 'categorizer' in app.ml_service.models
            # Try to get accuracy from model metadata if available
            if model_loaded and hasattr(app.ml_service.models.get('categorizer'), 'accuracy_'):
                model_accuracy = app.ml_service.models['categorizer'].accuracy_

        return jsonify({
            'categorized_items': categorized_items,
            'total_categories': total_categories,
            'model_loaded': model_loaded,
            'model_accuracy': model_accuracy
        })
    except Exception as e:
        app.logger.error(f"ML stats error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ml/train-categorizer', methods=['POST'])
@login_required
@requires_permission('manage_settings')
@limiter.limit("3 per hour")  # Rate limit for resource-intensive ML training
def train_categorizer():
    """Train the categorizer model"""
    try:
        if not hasattr(app, 'ml_service') or not app.ml_service:
            return jsonify({
                'success': False,
                'error': 'ML Service ist nicht aktiviert'
            }), 400

        # Get all items for training
        items = Item.query.filter(Item.category_id.isnot(None)).all()

        if len(items) < 10:
            return jsonify({
                'success': False,
                'error': f'Nicht genug Trainingsdaten. {len(items)} Gegenstände gefunden, mindestens 10 benötigt.'
            }), 400

        # Train the model
        result = app.ml_service.train_categorizer(items)

        if result.get('success'):
            app.logger.info(f"Categorizer model trained successfully with {result['accuracy']:.2%} accuracy")

        return jsonify(result)

    except Exception as e:
        app.logger.error(f"ML training error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ─── Dell Warranty Sync Routes ─────────────────────────────────────────────

@app.route('/api/dell/sync-warranty/<int:item_id>', methods=['POST'])
@login_required
@requires_permission('manage_items')
def dell_sync_warranty_single(item_id):
    """Synchronisiert Dell Garantiedaten für ein einzelnes Gerät"""
    try:
        item = Item.query.get_or_404(item_id)
        success, message = sync_dell_warranty(item)

        if success:
            log_action('dell_sync', 'item', item_id)
            return jsonify({'success': True, 'message': message, 'data': {
                'warranty_status': item.dell_warranty_status,
                'warranty_end_date': item.dell_warranty_end_date.isoformat() if item.dell_warranty_end_date else None,
                'contract_type': item.dell_contract_type,
                'last_sync': item.dell_last_sync.isoformat() if item.dell_last_sync else None
            }})
        else:
            return jsonify({'success': False, 'message': message}), 400

    except Exception as e:
        app.logger.error(f"Dell sync error for item {item_id}: {e}")
        return jsonify({'success': False, 'message': f'Fehler: {str(e)}'}), 500

@app.route('/api/dell/sync-all-warranties', methods=['POST'])
@login_required
@requires_permission('manage_items')
@limiter.limit("5 per hour")  # Rate limit for external API calls
def dell_sync_all_warranties():
    """Synchronisiert Dell Garantiedaten für alle Geräte mit Dell Service Tag"""
    try:
        items = Item.query.filter(Item.dell_service_tag.isnot(None)).all()

        if not items:
            return jsonify({'success': False, 'message': 'Keine Geräte mit Dell Service Tag gefunden'}), 404

        results = {'success': 0, 'failed': 0, 'total': len(items), 'errors': []}

        for item in items:
            success, message = sync_dell_warranty(item)
            if success:
                results['success'] += 1
            else:
                results['failed'] += 1
                results['errors'].append(f"{item.name} ({item.dell_service_tag}): {message}")

        log_action('dell_sync_all', 'item')

        return jsonify({
            'success': True,
            'message': f'{results["success"]} von {results["total"]} Geräten erfolgreich synchronisiert',
            'data': results
        })

    except Exception as e:
        app.logger.error(f"Dell bulk sync error: {e}")
        return jsonify({'success': False, 'message': f'Fehler: {str(e)}'}), 500

def setup():
    with app.app_context():
        # Tabellen anlegen
        db.create_all()

        # 1) Permissions anlegen
        perms = {
            'manage_users':        'Benutzerkonten verwalten',
            'manage_roles':        'Rollen & Berechtigungen verwalten',
            'manage_settings':     'System-Einstellungen verwalten',
            'view_dashboard':      'Dashboard anzeigen',
            'view_items':          'Gegenstände sehen',
            'view_loans':          'Ausleihen sehen',
            'view_tickets':        'Tickets sehen',
            'create_tickets':      'Tickets erstellen',
            'view_own_tickets':    'Eigene Tickets einsehen',
            'view_own_loans':      'Eigene Leihgaben einsehen',
            'answer_tickets':      'Tickets beantworten',
            'change_ticket_status':'Ticket-Status ändern',
            'view_admin_tickets':  'Admin-Tickets einsehen',
            'answer_admin_tickets':'Admin-Tickets beantworten',
            'manage_items':        'Items anlegen/bearbeiten'
        }
        for name, desc in perms.items():
            if not db.session.get(Permission, name):
                db.session.add(Permission(name=name, description=desc))

        # 2) Standard-Rollen anlegen
        for rn in ['admin', 'verwaltung', 'mitarbeiter', 'kunde']:
            if not db.session.get(Role, rn):
                db.session.add(Role(name=rn))

        db.session.commit()

        # 3) Default-Permissions zuweisen  ← Hier einsetzen
        admin_role = db.session.get(Role, 'admin')
        admin_role.permissions = Permission.query.all()

        verwaltung_role = db.session.get(Role, 'verwaltung')
        verwaltung_role.permissions = Permission.query.filter(
            Permission.name.in_([
                'view_dashboard',
                'view_items',
                'view_loans',
                'view_tickets',
                'view_borrowers',
                'answer_tickets',
                'change_ticket_status',
                'view_admin_tickets',
                'answer_admin_tickets',
                'manage_users',
                'manage_items'
            ])
        ).all()

        mitarbeiter_role = db.session.get(Role, 'mitarbeiter')
        mitarbeiter_role.permissions = Permission.query.filter(
            Permission.name.in_([
                'view_items',
                'view_loans',
                'create_tickets',
                'view_own_tickets',
                'view_own_loans'
            ])
        ).all()

        kunde_role = db.session.get(Role, 'kunde')
        kunde_role.permissions = Permission.query.filter(
            Permission.name.in_([
                'create_tickets',
                'view_own_tickets',
                'view_own_loans',
                'answer_tickets'
            ])
        ).all()

        db.session.commit()

        # 4) Default-Settings aus .env in DB anlegen
        defaults = {
            'MAIL_SERVER':   os.environ.get('MAIL_SERVER', ''),
            'MAIL_PORT':     os.environ.get('MAIL_PORT', '587'),
            'MAIL_USE_TLS':  os.environ.get('MAIL_USE_TLS', 'false'),
            'MAIL_USERNAME': os.environ.get('MAIL_USERNAME', ''),
            'MAIL_PASSWORD': os.environ.get('MAIL_PASSWORD', ''),
            'ADMINS':        os.environ.get('ADMIN_EMAIL', '').replace(',', ';'),
            'PRINTER_IP':    os.environ.get('PRINTER_IP', ''),
            'PRINTER_PORT':  os.environ.get('PRINTER_PORT', '9100'),
            'DELL_CLIENT_ID':     os.environ.get('DELL_CLIENT_ID',''),
            'DELL_CLIENT_SECRET': os.environ.get('DELL_CLIENT_SECRET',''),
        }
        for key, val in defaults.items():
            setting = db.session.get(Setting, key)
            if not setting:
                db.session.add(Setting(key=key, value=str(val)))

        db.session.commit()

        # 5) Check if admin exists, otherwise show setup instructions
        if not User.query.filter_by(role = UserRole.ADMIN).first():
            print("\n" + "="*70)
            print("  ⚠️  KEIN ADMIN-ACCOUNT GEFUNDEN")
            print("="*70)
            print("\nBitte erstellen Sie einen Admin-Account über die Python-Shell:")
            print("\n  python")
            print("  >>> from app import app, db, User")
            print("  >>> with app.app_context():")
            print("  ...     admin = User(username='admin', email='admin@example.com', role = UserRole.ADMIN)")
            print("  ...     admin.set_password('IhrSicheresPasswort123!')")
            print("  ...     db.session.add(admin)")
            print("  ...     db.session.commit()")
            print("="*70 + "\n")

# ─── Error Handlers ────────────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('errors/403.html'), 403

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    app.logger.error(f'Server Error: {error}')
    return render_template('errors/500.html'), 500

@app.errorhandler(429)
def ratelimit_handler(error):
    return render_template('errors/429.html'), 429

@app.route('/health')
def health_check():
    """Health check endpoint for Docker/Kubernetes"""
    try:
        # Check database connection
        db.session.execute('SELECT 1')
        return jsonify({
            'status': 'healthy',
            'version': APP_VERSION,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 503

# ═══════════════════════════════════════════════════════════════════════════════
# NEUE ROUTES FÜR FEATURES
# ═══════════════════════════════════════════════════════════════════════════════

# ─── Feature 1: QR-Code System ────────────────────────────────────────────────

@app.route('/item/<int:item_id>/qr')
@login_required
def generate_item_qr(item_id):
    """Generiert QR-Code für ein Item"""
    item = Item.query.get_or_404(item_id)

    if not item.qr_code:
        base_url = request.url_root.rstrip('/')
        item.qr_code = generate_qr_code(item_id, base_url)
        db.session.commit()
        log_action('generate_qr', 'item', item_id)

    return jsonify({'qr_code': item.qr_code})

@app.route('/items/qr-codes/pdf')
@login_required
@requires_permission('manage_items')
def export_qr_codes_pdf():
    """Exportiert alle QR-Codes als PDF zum Ausdrucken"""
    items = Item.query.all()
    base_url = request.url_root.rstrip('/')

    # Generiere QR-Codes für Items die noch keinen haben
    for item in items:
        if not item.qr_code:
            item.qr_code = generate_qr_code(item.id, base_url)

    db.session.commit()

    return render_template('qr_codes_pdf.html', items=items)

@app.route('/scan/qr', methods=['GET', 'POST'])
@login_required
def scan_qr():
    """QR-Code Scanner Interface"""
    if request.method == 'POST':
        data = request.get_json()
        qr_data = data.get('qr_data')

        # Extrahiere Item-ID aus URL
        try:
            item_id = int(qr_data.split('/item/')[-1].split('?')[0])
            item = Item.query.get(item_id)

            if item:
                log_action('qr_scan', 'item', item_id)
                return jsonify({'success': True, 'item_id': item_id, 'item_name': item.name})
        except (ValueError, IndexError, AttributeError) as e:
            app.logger.warning(f"Invalid QR code data: {qr_data[:100]} - Error: {e}")

        return jsonify({'success': False, 'message': 'Ungültiger QR-Code'}), 400

    return render_template('scan/qr.html')

# ─── Feature 2: Wartungs-Tracker ──────────────────────────────────────────────

@app.route('/maintenance')
@login_required
@requires_permission('manage_items')
def maintenance_list():
    """Liste aller Wartungen"""
    status_filter = request.args.get('status', 'all')

    query = Maintenance.query.join(Item)

    if status_filter != 'all':
        query = query.filter(Maintenance.status == status_filter)

    maintenances = query.order_by(Maintenance.scheduled_date.asc()).all()

    return render_template('maintenance/list.html', maintenances=maintenances, status_filter=status_filter)

@app.route('/maintenance/add/<int:item_id>', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def add_maintenance(item_id):
    """Neue Wartung planen"""
    item = Item.query.get_or_404(item_id)

    if request.method == 'POST':
        maintenance = Maintenance(
            item_id=item_id,
            scheduled_date=datetime.strptime(request.form['scheduled_date'], '%Y-%m-%d').date(),
            maintenance_type=request.form['maintenance_type'],
            notes=request.form.get('notes'),
            performed_by_id=current_user.id if request.form.get('assign_to_me') else None
        )

        db.session.add(maintenance)
        db.session.commit()

        log_action('create_maintenance', 'maintenance', maintenance.id)
        flash('Wartung erfolgreich geplant', FlashCategory.SUCCESS)

        return redirect(url_for('item_detail', item_id=item_id))

    return render_template('maintenance/add.html', item=item)

@app.route('/maintenance/<int:maintenance_id>/complete', methods=['POST'])
@login_required
@requires_permission('manage_items')
def complete_maintenance(maintenance_id):
    """Wartung als abgeschlossen markieren"""
    maintenance = Maintenance.query.get_or_404(maintenance_id)

    maintenance.status = 'completed'
    maintenance.completed_date = date.today()
    maintenance.notes = request.form.get('notes', maintenance.notes)
    maintenance.cost = float(request.form.get('cost', 0))

    # Nächste Wartung planen wenn Intervall gesetzt ist
    item = maintenance.item
    if item.maintenance_interval_days:
        next_date = date.today() + timedelta(days=item.maintenance_interval_days)
        item.next_maintenance_date = next_date

        # Automatisch neue Wartung erstellen
        next_maintenance = Maintenance(
            item_id=item.id,
            scheduled_date=next_date,
            maintenance_type=maintenance.maintenance_type,
            status='scheduled'
        )
        db.session.add(next_maintenance)

    db.session.commit()
    log_action('complete_maintenance', 'maintenance', maintenance_id)

    flash('Wartung abgeschlossen', FlashCategory.SUCCESS)
    return redirect(url_for('maintenance_list'))

# ─── Feature 3: Reservierungssystem ───────────────────────────────────────────

@app.route('/reservations')
@login_required
def reservation_list():
    """Liste aller Reservierungen"""
    if current_user.role == UserRole.ADMIN:
        reservations = Reservation.query.order_by(Reservation.start_date.desc()).all()
    else:
        reservations = Reservation.query.filter_by(user_id=current_user.id).order_by(Reservation.start_date.desc()).all()

    return render_template('reservations/list.html', reservations=reservations)

@app.route('/reservations/calendar')
@login_required
def reservation_calendar():
    """Kalenderansicht aller Reservierungen"""
    try:
        item_id = request.args.get('item_id', type=int)

        query = Reservation.query.filter(Reservation.status.in_(['pending', 'confirmed']))

        if item_id:
            query = query.filter_by(item_id=item_id)

        reservations = query.all()

        # Konvertiere für FullCalendar
        events = []
        for res in reservations:
            try:
                # Sichere Zugriffe auf Relationen
                item_name = res.item.name if res.item else f"Item #{res.item_id}"
                user_name = res.user.username if res.user else f"User #{res.user_id}"

                events.append({
                    'id': res.id,
                    'title': f"{item_name} - {user_name}",
                    'start': res.start_date.isoformat(),
                    'end': res.end_date.isoformat(),
                    'color': 'green' if res.status == 'confirmed' else 'orange'
                })
            except Exception as e:
                app.logger.warning(f"Failed to process reservation {res.id}: {e}")
                continue

        items = Item.query.all()
        return render_template('reservations/calendar.html', events=events, items=items, selected_item_id=item_id)
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Reservation calendar error: {e}")
        flash('Fehler beim Laden des Kalenders', FlashCategory.DANGER)
        return render_template('reservations/calendar.html', events=[], items=[], selected_item_id=None)

@app.route('/reservations/add/<int:item_id>', methods=['GET', 'POST'])
@login_required
def add_reservation(item_id):
    """Neue Reservierung erstellen"""
    item = Item.query.get_or_404(item_id)

    if request.method == 'POST':
        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%dT%H:%M')
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%dT%H:%M')

        # Prüfe auf Konflikte
        has_conflict, conflicts = check_reservation_conflict(item_id, start_date, end_date)

        if has_conflict:
            flash(f'Konflikt: Item ist bereits für diesen Zeitraum reserviert von {conflicts[0].user.username}', FlashCategory.DANGER)
            return redirect(url_for('add_reservation', item_id=item_id))

        reservation = Reservation(
            item_id=item_id,
            user_id=current_user.id,
            start_date=start_date,
            end_date=end_date,
            notes=request.form.get('notes'),
            status='pending' if item.requires_approval else 'confirmed'
        )

        db.session.add(reservation)
        db.session.commit()

        # Benachrichtigung an Admins wenn Genehmigung erforderlich
        if item.requires_approval:
            admins = User.query.filter_by(role = UserRole.ADMIN).all()
            for admin in admins:
                create_notification(
                    admin.id,
                    'reservation_approval_needed',
                    'Neue Reservierung benötigt Genehmigung',
                    f'{current_user.username} möchte {item.name} reservieren',
                    f'/reservations/{reservation.id}'
                )
        else:
            create_notification(
                current_user.id,
                'reservation_confirmed',
                'Reservierung bestätigt',
                f'Ihre Reservierung für {item.name} wurde bestätigt',
                f'/reservations/{reservation.id}'
            )

        log_action('create_reservation', 'reservation', reservation.id)
        flash('Reservierung erstellt', FlashCategory.SUCCESS)

        return redirect(url_for('reservation_list'))

    return render_template('reservations/add.html', item=item)

@app.route('/reservations/<int:reservation_id>/cancel', methods=['POST'])
@login_required
def cancel_reservation(reservation_id):
    """Reservierung stornieren"""
    reservation = Reservation.query.get_or_404(reservation_id)

    # Nur eigene Reservierungen oder Admin darf stornieren
    if reservation.user_id != current_user.id and current_user.role != 'admin':
        abort(HTTPStatus.FORBIDDEN)

    reservation.status = 'cancelled'
    db.session.commit()

    log_action('cancel_reservation', 'reservation', reservation_id)
    flash('Reservierung storniert', FlashCategory.SUCCESS)

    return redirect(url_for('reservation_list'))

@app.route('/reservations/<int:reservation_id>/ical')
@login_required
def reservation_ical(reservation_id):
    """Exportiert Reservierung als iCal"""
    if not ICALENDAR_AVAILABLE:
        flash('iCal-Export ist nicht verfügbar. Bitte installieren Sie: pip install icalendar pytz', FlashCategory.WARNING)
        return redirect(url_for('reservation_list'))

    reservation = Reservation.query.get_or_404(reservation_id)

    # Nur eigene Reservierungen oder Admin
    if reservation.user_id != current_user.id and current_user.role != 'admin':
        abort(HTTPStatus.FORBIDDEN)

    ical_data = create_ical_event(
        title=f"Reservierung: {reservation.item.name}",
        start_date=reservation.start_date,
        end_date=reservation.end_date,
        description=reservation.notes or '',
        location=reservation.item.location or ''
    )

    if not ical_data:
        flash('iCal-Export fehlgeschlagen', FlashCategory.DANGER)
        return redirect(url_for('reservation_list'))

    return send_file(
        BytesIO(ical_data),
        mimetype='text/calendar',
        as_attachment=True,
        download_name=f'reservation_{reservation.id}.ics'
    )

# ─── Feature 5 & 6: Benachrichtigungen ───────────────────────────────────────

@app.route('/notifications')
@login_required
def notifications_list():
    """Liste aller Benachrichtigungen für den aktuellen User"""
    notifications = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).limit(50).all()

    # Markiere ungelesene als gelesen
    unread = Notification.query.filter_by(user_id=current_user.id, is_read=False).all()
    for notif in unread:
        notif.is_read = True
    db.session.commit()

    return render_template('notifications/list.html', notifications=notifications)

@app.route('/api/notifications/unread-count')
@login_required
def unread_notifications_count():
    """API Endpoint für Anzahl ungelesener Benachrichtigungen"""
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({'count': count})

@app.route('/api/notifications/mark-read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    """Benachrichtigung als gelesen markieren"""
    notification = Notification.query.get_or_404(notification_id)

    if notification.user_id != current_user.id:
        abort(HTTPStatus.FORBIDDEN)

    notification.is_read = True
    db.session.commit()

    return jsonify({'success': True})

# ─── Feature 6: Erweiterte Statistiken & Analytics ───────────────────────────

@app.route('/analytics')
@login_required
@requires_permission('manage_items')
def analytics_dashboard():
    """Erweiterte Analytics Dashboard"""
    try:
        # Gesamtstatistiken
        total_items = Item.query.count()
        borrowed_items = Item.query.filter_by(is_borrowed=True).count()
        defective_items = Item.query.filter_by(defective=True).count()
        total_users = User.query.count()

        # Top 10 meistgeliehene Items
        try:
            top_items = db.session.query(
                Item.name,
                func.count(Loan.id).label('loan_count')
            ).join(Loan).group_by(Item.id, Item.name).order_by(desc('loan_count')).limit(10).all()
        except Exception as e:
            app.logger.warning(f"Failed to get top items: {e}")
            top_items = []

        # Auslastung pro Kategorie
        try:
            category_stats = db.session.query(
                Item.category,
                func.count(Item.id).label('total'),
                func.sum(db.case((Item.is_borrowed == True, 1), else_=0)).label('borrowed')
            ).group_by(Item.category).all()
        except Exception as e:
            app.logger.warning(f"Failed to get category stats: {e}")
            category_stats = []

        # Wartungsstatistiken
        try:
            upcoming_maintenance = Maintenance.query.filter(
                Maintenance.status == 'scheduled',
                Maintenance.scheduled_date >= date.today()
            ).count()
        except Exception as e:
            app.logger.warning(f"Failed to get upcoming maintenance: {e}")
            upcoming_maintenance = 0

        try:
            overdue_maintenance = Maintenance.query.filter(
                Maintenance.status == 'overdue'
            ).count()
        except Exception as e:
            app.logger.warning(f"Failed to get overdue maintenance: {e}")
            overdue_maintenance = 0

        # Budget-Übersicht
        try:
            total_value = db.session.query(func.sum(Item.purchase_price)).scalar() or 0
            current_total_value = db.session.query(func.sum(Item.current_value)).scalar() or 0
        except Exception as e:
            app.logger.warning(f"Failed to get budget overview: {e}")
            total_value = 0
            current_total_value = 0

        # Monatliche Ausleihen (letzten 12 Monate) - SQLite/PostgreSQL kompatibel
        try:
            twelve_months_ago = date.today() - timedelta(days=365)
            # Verwende date_trunc für PostgreSQL/SQLite Kompatibilität
            monthly_loans = db.session.query(
                func.date_trunc('month', Loan.loan_date).label('month'),
                func.count(Loan.id).label('count')
            ).filter(Loan.loan_date >= twelve_months_ago).group_by('month').all()
        except Exception as e:
            app.logger.warning(f"Failed to get monthly loans: {e}")
            monthly_loans = []

        return render_template('analytics/dashboard.html',
            total_items=total_items,
            borrowed_items=borrowed_items,
            defective_items=defective_items,
            total_users=total_users,
            top_items=top_items,
            category_stats=category_stats,
            upcoming_maintenance=upcoming_maintenance,
            overdue_maintenance=overdue_maintenance,
            total_value=total_value,
            current_total_value=current_total_value,
            monthly_loans=monthly_loans
        )
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Analytics dashboard error: {e}")
        flash('Fehler beim Laden des Analytics Dashboards', FlashCategory.DANGER)
        return render_template('analytics/dashboard.html',
            total_items=0,
            borrowed_items=0,
            defective_items=0,
            total_users=0,
            top_items=[],
            category_stats=[],
            upcoming_maintenance=0,
            overdue_maintenance=0,
            total_value=0,
            current_total_value=0,
            monthly_loans=[]
        )

@app.route('/analytics/export')
@login_required
@requires_permission('manage_items')
def analytics_export():
    """Exportiert Analytics als Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        flash('openpyxl ist nicht installiert. Export nicht möglich.', FlashCategory.DANGER)
        return redirect(url_for('analytics_dashboard'))

    try:
        wb = Workbook()

        # Sheet 1: Übersicht
        ws1 = wb.active
        ws1.title = "Übersicht"
        ws1['A1'] = "Statistik"
        ws1['B1'] = "Wert"
        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)

        stats = [
            ("Gesamt Items", Item.query.count()),
            ("Verliehen", Item.query.filter_by(is_borrowed=True).count()),
            ("Defekt", Item.query.filter_by(defective=True).count()),
            ("Gesamtwert", db.session.query(func.sum(Item.purchase_price)).scalar() or 0),
        ]

        for idx, (stat, value) in enumerate(stats, start=2):
            ws1[f'A{idx}'] = stat
            ws1[f'B{idx}'] = value

        # Sheet 2: Top Items
        ws2 = wb.create_sheet("Top Items")
        ws2['A1'] = "Item"
        ws2['B1'] = "Ausleihungen"

        try:
            top_items = db.session.query(
                Item.name,
                func.count(Loan.id).label('count')
            ).join(Loan).group_by(Item.id, Item.name).order_by(desc('count')).limit(20).all()

            for idx, (name, count) in enumerate(top_items, start=2):
                ws2[f'A{idx}'] = name
                ws2[f'B{idx}'] = count
        except Exception as e:
            app.logger.warning(f"Failed to get top items for export: {e}")
            ws2['A2'] = "Keine Daten verfügbar"

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        app.logger.error(f"Analytics export error: {e}")
        flash('Fehler beim Exportieren der Analytics', FlashCategory.DANGER)
        return redirect(url_for('analytics_dashboard'))

# ─── Feature 7: Budget-Tracking ──────────────────────────────────────────────

@app.route('/budget')
@login_required
@requires_permission('manage_items')
def budget_overview():
    """Budget-Übersicht"""
    try:
        current_year = date.today().year

        budgets = BudgetEntry.query.filter_by(year=current_year).all()

        # Gesamtinventar-Wert
        total_purchase_value = db.session.query(func.sum(Item.purchase_price)).scalar() or 0
        total_current_value = db.session.query(func.sum(Item.current_value)).scalar() or 0

        # Wartungskosten dieses Jahr - Kompatibel mit SQLite und PostgreSQL
        year_start = date(current_year, 1, 1)
        year_end = date(current_year, 12, 31)

        try:
            maintenance_costs = db.session.query(func.sum(Maintenance.cost)).filter(
                Maintenance.completed_date >= year_start,
                Maintenance.completed_date <= year_end
            ).scalar() or 0
        except Exception as e:
            app.logger.warning(f"Failed to get maintenance costs: {e}")
            maintenance_costs = 0

        # Reparaturkosten - Kompatibel mit SQLite und PostgreSQL
        try:
            repair_costs = db.session.query(func.sum(RepairLog.cost)).filter(
                RepairLog.repair_date >= year_start,
                RepairLog.repair_date <= year_end
            ).scalar() or 0
        except Exception as e:
            app.logger.warning(f"Failed to get repair costs: {e}")
            repair_costs = 0

        return render_template('budget/overview.html',
            budgets=budgets,
            total_value=total_purchase_value,
            total_purchase_value=total_purchase_value,
            current_value=total_current_value,
            total_current_value=total_current_value,
            maintenance_costs=maintenance_costs,
            repair_costs=repair_costs,
            current_year=current_year
        )
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Budget overview error: {e}")
        flash('Fehler beim Laden der Budget-Übersicht', FlashCategory.DANGER)
        return render_template('budget/overview.html',
            budgets=[],
            total_value=0,
            total_purchase_value=0,
            current_value=0,
            total_current_value=0,
            maintenance_costs=0,
            repair_costs=0,
            current_year=date.today().year
        )

@app.route('/budget/add', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def add_budget():
    """Create a new budget entry.

    Handles both GET requests (display form) and POST requests (process form submission).
    Creates a new BudgetEntry with all required fields including name, category,
    amount, date range, department, and description.
    """
    form = BudgetForm()

    if form.validate_on_submit():
        # Extract year from start_date
        year = form.start_date.data.year if form.start_date.data else datetime.now().year

        budget = BudgetEntry(
            name=form.name.data,
            year=year,
            category=form.category.data,
            allocated_amount=float(form.amount.data),
            start_date=form.start_date.data,
            end_date=form.end_date.data,
            description=form.description.data,
            department=form.department.data,
            notes=form.description.data  # Keep legacy notes field in sync
        )

        db.session.add(budget)
        db.session.commit()

        log_action('create_budget', 'budget', budget.id)
        flash('Budget erfolgreich erstellt', FlashCategory.SUCCESS)

        return redirect(url_for('budget_overview'))

    return render_template('budget/add.html', form=form)

@app.route('/budget/<int:budget_id>/transaction/add', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def add_budget_transaction(budget_id):
    """Add a transaction to a budget entry."""
    budget = BudgetEntry.query.get_or_404(budget_id)
    form = BudgetTransactionForm()

    if form.validate_on_submit():
        transaction = BudgetTransaction(
            budget_id=budget_id,
            amount=float(form.amount.data),
            transaction_type=form.transaction_type.data,
            transaction_date=form.transaction_date.data,
            description=form.description.data,
            created_by=current_user.id
        )
        db.session.add(transaction)

        # Update budget spent_amount
        budget.spent_amount = budget.calculate_spent_amount()
        db.session.commit()

        log_action('add_budget_transaction', 'budget_transaction', transaction.id)
        flash('Transaktion erfolgreich hinzugefügt', FlashCategory.SUCCESS)
        return redirect(url_for('budget_detail', budget_id=budget_id))

    return render_template('budget/add_transaction.html', form=form, budget=budget)

@app.route('/budget/<int:budget_id>')
@login_required
@requires_permission('manage_items')
def budget_detail(budget_id):
    """Display budget detail with transaction history."""
    budget = BudgetEntry.query.get_or_404(budget_id)

    # Get all transactions for this budget
    transactions = BudgetTransaction.query.filter_by(budget_id=budget_id).order_by(
        BudgetTransaction.transaction_date.desc()
    ).all()

    # Calculate current spent amount
    spent_amount = budget.calculate_spent_amount()
    remaining_budget = budget.get_remaining_budget()

    return render_template('budget/detail.html',
        budget=budget,
        transactions=transactions,
        spent_amount=spent_amount,
        remaining_budget=remaining_budget
    )

# ─── Feature 8: Standort-Management ──────────────────────────────────────────

@app.route('/locations')
@login_required
@requires_permission('manage_items')
def location_list():
    """Liste aller Standorte"""
    try:
        locations = Location.query.filter(Location.parent_id.is_(None)).all()
    except Exception as e:
        app.logger.error(f"Failed to load locations: {e}")
        flash('Fehler beim Laden der Standorte', FlashCategory.DANGER)
        locations = []

    return render_template('locations/list.html', locations=locations)

# OLD Location routes - replaced with new /admin/locations routes above
# @app.route('/locations/add', methods=['GET', 'POST'])
# @login_required
# @requires_permission('manage_items')
# def add_location_old():
#     """Neuen Standort hinzufügen"""
#     if request.method == 'POST':
#         try:
#             location = Location(
#                 name=request.form['name'],
#                 description=request.form.get('description'),
#                 parent_id=int(request.form['parent_id']) if request.form.get('parent_id') else None
#             )
#
#             db.session.add(location)
#             db.session.commit()
#
#             log_action('create_location', 'location', location.id)
#             flash('Standort erstellt', FlashCategory.SUCCESS)
#
#             return redirect(url_for('location_list'))
#         except Exception as e:
#             db.session.rollback()
#             app.logger.error(f"Failed to create location: {e}")
#             flash(f'Fehler beim Erstellen des Standorts: {str(e)}', FlashCategory.DANGER)
#
#     # Alle Locations für Parent-Auswahl
#     try:
#         all_locations = Location.query.all()
#     except Exception as e:
#         app.logger.warning(f"Failed to load locations: {e}")
#         all_locations = []
#
#     return render_template('locations/add.html', all_locations=all_locations)

@app.route('/locations/<int:location_id>')
@login_required
@requires_permission('manage_items')
def location_detail(location_id):
    """Standort-Details mit hierarchie"""
    location = Location.query.get_or_404(location_id)
    items = Item.query.filter_by(location_id=location_id).all()

    return render_template('locations/detail.html', location=location, items=items)

# ─── Feature 10: Smart Suggestions ───────────────────────────────────────────

@app.route('/api/suggest-category', methods=['POST'])
@login_required
def suggest_category():
    """Schlägt automatisch Kategorie vor"""
    data = request.get_json()
    item_name = data.get('name', '')

    suggested_category = auto_categorize_item(item_name)

    return jsonify({'suggested_category': suggested_category})

@app.route('/api/similar-items/<int:item_id>')
@login_required
def get_similar_items(item_id):
    """Findet ähnliche Items"""
    item = Item.query.get_or_404(item_id)

    similar_items = find_similar_items(item.name, item.category, limit=5)

    return jsonify({
        'similar_items': [{
            'id': i.id,
            'name': i.name,
            'category': i.category,
            'is_borrowed': i.is_borrowed
        } for i in similar_items]
    })

# ─── Feature 11: Bulk-Operationen ────────────────────────────────────────────
# Note: bulk-import is now handled by routes around line 1904 with improved UI

@app.route('/items/bulk-edit', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def bulk_edit_items():
    """Bulk-Bearbeitung von Items"""
    if request.method == 'POST':
        item_ids = request.form.getlist('item_ids')
        action = request.form.get('action')

        if not item_ids:
            flash('Keine Items ausgewählt', FlashCategory.WARNING)
            return redirect(url_for('items'))

        items = Item.query.filter(Item.id.in_([int(i) for i in item_ids])).all()

        if action == 'change_location':
            new_location_id = int(request.form['location_id']) if request.form.get('location_id') else None
            for item in items:
                item.location_id = new_location_id

        elif action == 'change_category':
            new_category = request.form['category']
            for item in items:
                item.category = new_category

        elif action == 'assign_team':
            team_id = int(request.form['team_id']) if request.form.get('team_id') else None
            for item in items:
                item.team_id = team_id

        elif action == 'mark_defective':
            for item in items:
                item.defective = True

        db.session.commit()
        log_action('bulk_edit', 'item', None, {'count': len(items), 'action': action})

        flash(f'{len(items)} Items erfolgreich aktualisiert', FlashCategory.SUCCESS)
        return redirect(url_for('items'))

    items = Item.query.all()
    locations = Location.query.all()
    teams = Team.query.all()

    return render_template('items/bulk_edit.html', items=items, locations=locations, teams=teams)

# ─── Feature 12: Defekt-Meldung ──────────────────────────────────────────────

@app.route('/defects')
@login_required
@requires_permission('manage_items')
def defect_list():
    """Liste aller Defektmeldungen"""
    status_filter = request.args.get('status', 'all')

    query = DefectReport.query.join(Item)

    if status_filter != 'all':
        query = query.filter(DefectReport.status == status_filter)

    defects = query.order_by(DefectReport.created_at.desc()).all()

    return render_template('defects/list.html', defects=defects, status_filter=status_filter)

@app.route('/defects/report/<int:item_id>', methods=['GET', 'POST'])
@login_required
def report_defect(item_id):
    """Defekt melden"""
    item = Item.query.get_or_404(item_id)

    if request.method == 'POST':
        defect = DefectReport(
            item_id=item_id,
            reported_by_id=current_user.id,
            description=request.form['description'],
            severity=request.form.get('severity', 'medium')
        )

        # Markiere Item als defekt
        item.defective = True

        db.session.add(defect)
        db.session.commit()

        # Benachrichtige Admins
        admins = User.query.filter_by(role = UserRole.ADMIN).all()
        for admin in admins:
            create_notification(
                admin.id,
                'defect_reported',
                'Neuer Defekt gemeldet',
                f'{current_user.username} hat einen Defekt bei {item.name} gemeldet',
                f'/defects/{defect.id}'
            )

        log_action('report_defect', 'defect', defect.id)
        flash('Defekt gemeldet', FlashCategory.SUCCESS)

        return redirect(url_for('item_detail', item_id=item_id))

    return render_template('defects/report.html', item=item)

@app.route('/defects/<int:defect_id>/repair', methods=['POST'])
@login_required
@requires_permission('manage_items')
def log_repair(defect_id):
    """Reparatur protokollieren"""
    defect = DefectReport.query.get_or_404(defect_id)

    repair_log = RepairLog(
        defect_report_id=defect_id,
        repaired_by_id=current_user.id,
        action_taken=request.form['action_taken'],
        parts_replaced=request.form.get('parts_replaced'),
        cost=float(request.form.get('cost', 0))
    )

    defect.status = 'fixed'
    defect.item.defective = False

    db.session.add(repair_log)
    db.session.commit()

    # Benachrichtige Melder
    create_notification(
        defect.reported_by_id,
        'defect_fixed',
        'Defekt behoben',
        f'Der von Ihnen gemeldete Defekt bei {defect.item.name} wurde behoben',
        f'/item/{defect.item_id}'
    )

    log_action('repair_item', 'repair_log', repair_log.id)
    flash('Reparatur protokolliert', FlashCategory.SUCCESS)

    return redirect(url_for('defect_list'))

# ─── Feature 13: Kommentar & Rating ──────────────────────────────────────────

@app.route('/item/<int:item_id>/comment', methods=['POST'])
@login_required
def add_comment(item_id):
    """Kommentar zu Item hinzufügen"""
    item = Item.query.get_or_404(item_id)

    comment = Comment(
        item_id=item_id,
        user_id=current_user.id,
        content=request.form['content'],
        is_internal=request.form.get('is_internal') == 'on'
    )

    db.session.add(comment)
    db.session.commit()

    log_action('add_comment', 'comment', comment.id)
    flash('Kommentar hinzugefügt', FlashCategory.SUCCESS)

    return redirect(url_for('item_detail', item_id=item_id))

@app.route('/item/<int:item_id>/rate', methods=['POST'])
@login_required
def rate_item(item_id):
    """Item bewerten"""
    item = Item.query.get_or_404(item_id)

    # Prüfe ob User schon bewertet hat
    existing_rating = Rating.query.filter_by(item_id=item_id, user_id=current_user.id).first()

    stars = int(request.form['stars'])

    if existing_rating:
        existing_rating.stars = stars
        existing_rating.review = request.form.get('review')
    else:
        rating = Rating(
            item_id=item_id,
            user_id=current_user.id,
            stars=stars,
            review=request.form.get('review')
        )
        db.session.add(rating)

    db.session.commit()

    # Aktualisiere Durchschnittsbewertung
    avg_rating = db.session.query(func.avg(Rating.stars)).filter_by(item_id=item_id).scalar()
    rating_count = Rating.query.filter_by(item_id=item_id).count()

    item.average_rating = round(avg_rating, 2)
    item.rating_count = rating_count

    db.session.commit()

    log_action('rate_item', 'rating', item_id)
    flash('Bewertung gespeichert', FlashCategory.SUCCESS)

    return redirect(url_for('item_detail', item_id=item_id))

# ─── Feature 14: Genehmigungs-Workflow ───────────────────────────────────────

@app.route('/approvals')
@login_required
def approval_requests():
    """Liste aller Genehmigungsanfragen"""
    if current_user.role == UserRole.ADMIN:
        requests = ApprovalRequest.query.filter_by(status='pending').order_by(ApprovalRequest.created_at.desc()).all()
    else:
        requests = ApprovalRequest.query.filter_by(requester_id=current_user.id).order_by(ApprovalRequest.created_at.desc()).all()

    return render_template('approvals/list.html', requests=requests)

@app.route('/approvals/<int:request_id>/approve', methods=['POST'])
@login_required
@requires_permission('manage_items')
def approve_request(request_id):
    """Genehmigungsanfrage genehmigen"""
    approval_request = ApprovalRequest.query.get_or_404(request_id)

    approval_request.status = 'approved'
    approval_request.approver_id = current_user.id

    # Erstelle automatisch eine Ausleihe
    loan = Loan(
        item_id=approval_request.item_id,
        borrower_id=approval_request.requester_id,
        due_date=approval_request.end_date
    )

    approval_request.item.is_borrowed = True

    db.session.add(loan)
    db.session.commit()

    # Benachrichtige Antragsteller
    create_notification(
        approval_request.requester_id,
        'approval_granted',
        'Genehmigung erteilt',
        f'Ihre Anfrage für {approval_request.item.name} wurde genehmigt',
        f'/loans'
    )

    log_action('approve_request', 'approval_request', request_id)
    flash('Anfrage genehmigt', FlashCategory.SUCCESS)

    return redirect(url_for('approval_requests'))

@app.route('/approvals/<int:request_id>/reject', methods=['POST'])
@login_required
@requires_permission('manage_items')
def reject_request(request_id):
    """Genehmigungsanfrage ablehnen"""
    approval_request = ApprovalRequest.query.get_or_404(request_id)

    approval_request.status = 'rejected'
    approval_request.approver_id = current_user.id
    approval_request.rejection_reason = request.form.get('reason')

    db.session.commit()

    # Benachrichtige Antragsteller
    create_notification(
        approval_request.requester_id,
        'approval_rejected',
        'Anfrage abgelehnt',
        f'Ihre Anfrage für {approval_request.item.name} wurde abgelehnt. Grund: {approval_request.rejection_reason}',
        f'/approvals'
    )

    log_action('reject_request', 'approval_request', request_id)
    flash('Anfrage abgelehnt', FlashCategory.SUCCESS)

    return redirect(url_for('approval_requests'))

# ─── Feature 15: Team-Verwaltung ─────────────────────────────────────────────

@app.route('/teams')
@login_required
@requires_permission('manage_items')
def team_list():
    """Liste aller Teams"""
    teams = Team.query.all()
    return render_template('teams/list.html', teams=teams)

@app.route('/teams/add', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def add_team():
    """Create a new team with full configuration.

    Handles both GET requests (display form) and POST requests (process form submission).
    Creates a new Team with all required fields including name, department, contact
    information, permissions, and status. Members are handled separately after creation.
    """
    form = TeamForm()

    if form.validate_on_submit():
        team = Team(
            name=form.name.data,
            description=form.description.data,
            department=form.department.data,
            team_lead=form.team_lead.data,
            email=form.email.data,
            phone=form.phone.data,
            can_view_all=form.can_view_all.data,
            can_edit_all=form.can_edit_all.data,
            can_delete_all=form.can_delete_all.data,
            can_manage_team=form.can_manage_team.data,
            is_active=form.is_active.data
        )

        db.session.add(team)
        db.session.commit()

        log_action('create_team', 'team', team.id)
        flash('Team erfolgreich erstellt', FlashCategory.SUCCESS)

        return redirect(url_for('team_list'))

    return render_template('teams/add.html', form=form)

@app.route('/teams/<int:team_id>')
@login_required
def team_detail(team_id):
    """Team-Details"""
    team = Team.query.get_or_404(team_id)
    return render_template('teams/detail.html', team=team)

@app.route('/teams/<int:team_id>/add-member', methods=['POST'])
@login_required
@requires_permission('manage_items')
def add_team_member(team_id):
    """Mitglied zu Team hinzufügen"""
    team = Team.query.get_or_404(team_id)

    user_id = int(request.form['user_id'])
    role = request.form.get('role', 'member')

    # Prüfe ob schon Mitglied
    existing = TeamMember.query.filter_by(team_id=team_id, user_id=user_id).first()
    if existing:
        flash('Benutzer ist bereits Mitglied dieses Teams', FlashCategory.WARNING)
        return redirect(url_for('team_detail', team_id=team_id))

    member = TeamMember(
        team_id=team_id,
        user_id=user_id,
        role=role
    )

    db.session.add(member)
    db.session.commit()

    log_action('add_team_member', 'team_member', member.id)
    flash('Mitglied hinzugefügt', FlashCategory.SUCCESS)

    return redirect(url_for('team_detail', team_id=team_id))

# ─── Feature 18: Dark Mode & Theme ───────────────────────────────────────────

@app.route('/settings/theme', methods=['POST'])
@login_required
def update_theme():
    """Theme-Präferenz aktualisieren"""
    theme = request.form.get('theme', 'auto')

    if theme not in ['light', 'dark', 'auto']:
        theme = 'auto'

    current_user.theme_preference = theme
    db.session.commit()

    return jsonify({'success': True, 'theme': theme})

@app.context_processor
def inject_theme():
    """Macht Theme für alle Templates verfügbar"""
    if current_user.is_authenticated:
        return {'user_theme': current_user.theme_preference}
    return {'user_theme': 'auto'}

# ─── Feature 19: Erweiterte Suche & Filter ───────────────────────────────────

@app.route('/search')
@login_required
def advanced_search():
    """Erweiterte Suchseite"""
    query = Item.query

    # Textsuche
    search_term = request.args.get('q')
    if search_term:
        query = query.filter(
            or_(
                Item.name.ilike(f'%{search_term}%'),
                Item.serial.ilike(f'%{search_term}%'),
                Item.note.ilike(f'%{search_term}%')
            )
        )

    # Filter
    category = request.args.get('category')
    if category:
        query = query.filter_by(category=category)

    location_id = request.args.get('location_id', type=int)
    if location_id:
        query = query.filter_by(location_id=location_id)

    team_id = request.args.get('team_id', type=int)
    if team_id:
        query = query.filter_by(team_id=team_id)

    is_borrowed = request.args.get('is_borrowed')
    if is_borrowed == 'yes':
        query = query.filter_by(is_borrowed=True)
    elif is_borrowed == 'no':
        query = query.filter_by(is_borrowed=False)

    defective = request.args.get('defective')
    if defective == 'yes':
        query = query.filter_by(defective=True)
    elif defective == 'no':
        query = query.filter_by(defective=False)

    # Sortierung
    sort_by = request.args.get('sort_by', 'name')
    if sort_by == 'name':
        query = query.order_by(Item.name)
    elif sort_by == 'serial':
        query = query.order_by(Item.serial)
    elif sort_by == 'created_at':
        query = query.order_by(Item.created_at.desc())

    items = query.all()

    locations = Location.query.all()
    teams = Team.query.all()

    return render_template('search/advanced.html', items=items, locations=locations, teams=teams)

@app.route('/filters/save', methods=['POST'])
@login_required
def save_filter():
    """Speichert einen Filter"""
    data = request.get_json()

    saved_filter = SavedFilter(
        user_id=current_user.id,
        name=data['name'],
        filter_type=data['filter_type'],
        filter_params=json.dumps(data['params']),
        is_default=data.get('is_default', False)
    )

    # Nur ein Default-Filter pro Typ
    if saved_filter.is_default:
        SavedFilter.query.filter_by(
            user_id=current_user.id,
            filter_type=saved_filter.filter_type,
            is_default=True
        ).update({'is_default': False})

    db.session.add(saved_filter)
    db.session.commit()

    log_action('save_filter', 'saved_filter', saved_filter.id)

    return jsonify({'success': True, 'filter_id': saved_filter.id})

@app.route('/filters/my')
@login_required
def my_filters():
    """Zeigt gespeicherte Filter des Users"""
    try:
        filters = SavedFilter.query.filter_by(user_id=current_user.id).all()
    except Exception as e:
        app.logger.error(f"Failed to load filters: {e}")
        flash('Fehler beim Laden der Filter', FlashCategory.DANGER)
        filters = []

    return render_template('filters/list.html', filters=filters)

# ─── Feature 20: Drag & Drop Uploads ─────────────────────────────────────────

@app.route('/api/upload/image', methods=['POST'])
@login_required
@requires_permission('manage_items')
@limiter.limit("30 per minute")  # Rate limit for image uploads
def upload_image():
    """Drag & Drop Image Upload"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'Keine Datei'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'message': 'Keine Datei ausgewählt'}), 400

    # Validiere Bildformat
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    if not any(file.filename.lower().endswith(ext) for ext in allowed_extensions):
        return jsonify({'success': False, 'message': 'Ungültiges Dateiformat'}), 400

    # Speichere Datei
    filename = secure_filename(file.filename)
    unique_filename = f"{uuid.uuid4().hex}_{filename}"

    upload_folder = os.path.join(app.root_path, 'static', 'uploads')
    os.makedirs(upload_folder, exist_ok=True)

    filepath = os.path.join(upload_folder, unique_filename)
    file.save(filepath)

    # Optional: Bild komprimieren
    try:
        img = Image.open(filepath)
        img.thumbnail((1920, 1920), Image.Resampling.LANCZOS)
        img.save(filepath, optimize=True, quality=85)
    except Exception as e:
        app.logger.warning(f"Image optimization failed: {e}")

    relative_path = f"/static/uploads/{unique_filename}"

    return jsonify({
        'success': True,
        'path': relative_path,
        'filename': unique_filename
    })

@app.route('/api/items/<int:item_id>/update-image', methods=['POST'])
@login_required
@requires_permission('manage_items')
def update_item_image(item_id):
    """Aktualisiert Item-Bild"""
    item = Item.query.get_or_404(item_id)

    data = request.get_json()
    image_path = data.get('image_path')

    item.image_path = image_path
    db.session.commit()

    log_action('update_item_image', 'item', item_id)

    return jsonify({'success': True})

# ─── Feature 21: Audit Log ───────────────────────────────────────────────────

@app.route('/audit-log')
@login_required
@requires_permission('manage_items')
def audit_log():
    """Audit Log Übersicht"""
    page = request.args.get('page', 1, type=int)
    per_page = 50

    action_filter = request.args.get('action')
    user_id_filter = request.args.get('user_id', type=int)

    query = AuditLog.query

    if action_filter:
        query = query.filter_by(action=action_filter)

    if user_id_filter:
        query = query.filter_by(user_id=user_id_filter)

    pagination = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    users = User.query.all()

    # Unique actions für Filter
    unique_actions = db.session.query(AuditLog.action).distinct().all()
    actions = [a[0] for a in unique_actions]

    return render_template('audit/log.html',
        pagination=pagination,
        users=users,
        actions=actions,
        action_filter=action_filter,
        user_id_filter=user_id_filter
    )

@app.route('/audit-log/export')
@login_required
@requires_permission('manage_items')
def export_audit_log():
    """Exportiert Audit Log als CSV"""
    import csv
    from io import StringIO

    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(['Timestamp', 'User', 'Action', 'Entity Type', 'Entity ID', 'IP Address'])

    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(10000).all()

    for log in logs:
        writer.writerow([
            log.timestamp.isoformat(),
            log.user.username if log.user else 'System',
            log.action,
            log.entity_type,
            log.entity_id,
            log.ip_address
        ])

    return send_file(
        BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'audit_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    )

# ─── Feature 22: Granulare Berechtigungen ────────────────────────────────────

@app.route('/permissions')
@login_required
@requires_permission('manage_items')
def permission_list():
    """Liste aller Berechtigungen"""
    try:
        permissions = UserPermission.query.all()
    except Exception as e:
        app.logger.error(f"Failed to load permissions: {e}")
        flash('Fehler beim Laden der Berechtigungen', FlashCategory.DANGER)
        permissions = []

    return render_template('permissions/list.html', permissions=permissions)

@app.route('/permissions/add', methods=['GET', 'POST'])
@login_required
@requires_permission('manage_items')
def add_permission():
    """Neue Berechtigung erstellen"""
    if request.method == 'POST':
        try:
            permission = UserPermission(
                user_id=int(request.form['user_id']) if request.form.get('user_id') else None,
                role=request.form.get('role'),
                resource_type=request.form['resource_type'],
                resource_id=int(request.form['resource_id']) if request.form.get('resource_id') else None,
                action=request.form['action'],
                granted=request.form.get('granted') == 'on',
                expires_at=datetime.strptime(request.form['expires_at'], '%Y-%m-%d') if request.form.get('expires_at') else None
            )

            db.session.add(permission)
            db.session.commit()

            log_action('create_permission', 'permission', permission.id)
            flash('Berechtigung erstellt', FlashCategory.SUCCESS)

            return redirect(url_for('permission_list'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Failed to create permission: {e}")
            flash(f'Fehler beim Erstellen der Berechtigung: {str(e)}', FlashCategory.DANGER)

    try:
        users = User.query.all()
    except Exception as e:
        app.logger.warning(f"Failed to load users: {e}")
        users = []

    return render_template('permissions/add.html', users=users)

# ─── Feature 23 & 24: Barcode & RFID Scanner ─────────────────────────────────

@app.route('/scan/barcode', methods=['GET', 'POST'])
@login_required
def scan_barcode():
    """Barcode Scanner Interface"""
    if request.method == 'POST':
        data = request.get_json()
        barcode = data.get('barcode')

        result = process_barcode_scan(barcode)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 404

    return render_template('scan/barcode.html')

@app.route('/scan/rfid', methods=['GET', 'POST'])
@login_required
def scan_rfid():
    """RFID Scanner Interface"""
    if request.method == 'POST':
        data = request.get_json()
        rfid_tag = data.get('rfid_tag')

        result = process_rfid_scan(rfid_tag)

        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 404

    return render_template('scan/rfid.html')

@app.route('/scan/bluetooth')
@login_required
def scan_bluetooth():
    """Bluetooth Scanner Interface mit Web Bluetooth API"""
    return render_template('scan/bluetooth.html')

@app.route('/api/items/<int:item_id>/rfid', methods=['POST'])
@login_required
@requires_permission('manage_items')
def assign_rfid_tag(item_id):
    """RFID-Tag einem Item zuweisen"""
    item = Item.query.get_or_404(item_id)

    data = request.get_json()
    rfid_tag_id = data.get('rfid_tag_id')

    # Prüfe ob Tag schon verwendet wird
    existing = Item.query.filter_by(rfid_tag_id=rfid_tag_id).first()
    if existing and existing.id != item_id:
        return jsonify({'success': False, 'message': 'RFID-Tag bereits verwendet'}), 400

    item.rfid_tag_id = rfid_tag_id
    db.session.commit()

    log_action('assign_rfid', 'item', item_id, {'rfid_tag': rfid_tag_id})

    return jsonify({'success': True})

# ─── Feature 4: PWA Support ──────────────────────────────────────────────────

@app.route('/manifest.json')
def manifest():
    """PWA Manifest"""
    return jsonify({
        "name": "InventoryApp",
        "short_name": "Inventory",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#0d6efd",
        "description": "Professional Inventory Management System",
        "icons": [
            {
                "src": "/static/icon-192.png",
                "sizes": "192x192",
                "type": "image/png"
            },
            {
                "src": "/static/icon-512.png",
                "sizes": "512x512",
                "type": "image/png"
            }
        ]
    })

@app.route('/service-worker.js')
def service_worker():
    """Service Worker für Offline-Funktionalität"""
    return send_file('static/service-worker.js', mimetype='application/javascript')
# Auth route 'LDAP login' migrated to routes/auth.py

# ═══════════════════════════════════════════════════════════════════════════
# ─── v2.3.0 FEATURES INITIALIZATION ────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════

# Initialize SocketIO Events
if SOCKETIO_AVAILABLE and socketio:
    try:
        from socketio_events import register_socketio_events
        register_socketio_events(socketio, app)
        app.logger.info("✅ SocketIO events registered")
    except Exception as e:
        app.logger.error(f"SocketIO events registration failed: {e}")

# Initialize API Routes (GraphQL, Analytics, Reports, Mobile API)
try:
    from api_routes import init_api_routes
    init_api_routes(app, db, GRAPHQL_AVAILABLE, PLOTLY_AVAILABLE, socketio)
    app.logger.info("✅ API routes initialized")
except Exception as e:
    app.logger.error(f"API routes initialization failed: {e}")
    import traceback
    app.logger.error(traceback.format_exc())

# Initialize Integration Manager with Lazy Loading
try:
    from integrations import IntegrationManager
    integration_manager = IntegrationManager(app, db, mail)
    app.integration_manager = integration_manager
    app.logger.info("✅ Integration Manager initialized (lazy loading enabled)")
except Exception as e:
    app.logger.error(f"Integration Manager initialization failed: {e}")
    app.integration_manager = None

# Initialize Printer Manager (USB + Network Printers)
try:
    from printer_manager import init_printer_routes
    # Pass None for requires_permission as the module can use @login_required instead
    init_printer_routes(app, db, None)
    app.logger.info("✅ Printer Manager initialized (USB + Network support)")
except Exception as e:
    app.logger.error(f"Printer Manager initialization failed: {e}")

# Initialize Label Generator (Barcodes + QR Codes)
try:
    from label_generator import init_label_routes
    # Pass None for requires_permission as the module can use @login_required instead
    init_label_routes(app, db, None)
    app.logger.info("✅ Label Generator initialized (Barcode + QR Code support)")
except Exception as e:
    app.logger.error(f"Label Generator initialization failed: {e}")

# Initialize ML Service (Machine Learning for Smart Features)
try:
    from ml_service import MLService
    ml_service = MLService(app)
    app.ml_service = ml_service
    app.logger.info("✅ ML Service initialized (Smart Categorization, Predictive Maintenance)")
except Exception as e:
    app.logger.error(f"ML Service initialization failed: {e}")
    app.ml_service = None

# ─── Log Management ─────────────────────────────────────────────────────────
try:
    from utils.log_manager import init_log_routes
    init_log_routes(app, db)
    app.logger.info("✅ Log Management initialized")
except Exception as e:
    app.logger.error(f"Log Management initialization failed: {e}")

# ─── Admin Routes Blueprint ────────────────────────────────────────────────
try:
    from routes.admin import init_admin_routes
    init_admin_routes(app, db)
    app.logger.info("✅ Admin routes blueprint initialized")
except Exception as e:
    app.logger.error(f"Admin routes blueprint initialization failed: {e}")
    import traceback
    app.logger.error(traceback.format_exc())

# ─── Auth Routes Blueprint ──────────────────────────────────────────────────
try:
    from routes.auth import init_auth_routes
    init_auth_routes(app, db, mail, limiter, OAUTH_AVAILABLE, oauth, SAML_AVAILABLE)
    app.logger.info("✅ Auth routes blueprint initialized")
except Exception as e:
    app.logger.error(f"Auth routes blueprint initialization failed: {e}")
    import traceback
    app.logger.error(traceback.format_exc())

# Initialize Enhanced Ticket System
# NOTE: Ticket routes are already defined in app.py, no need for separate initialization
# The ticket_system.py module is available for extended features if needed
app.logger.info("✅ Ticket System ready (routes defined in app.py)")

# Helper function to trigger integration events
def trigger_integration(event_type, data):
    """Trigger integration event if manager is available"""
    if hasattr(app, 'integration_manager') and app.integration_manager:
        try:
            app.integration_manager.trigger_event(event_type, data)
        except Exception as e:
            app.logger.error(f"Integration trigger failed: {e}")

# Helper function to broadcast WebSocket events
def broadcast_socketio(event_type, data):
    """Broadcast SocketIO event if available"""
    if SOCKETIO_AVAILABLE and socketio and hasattr(app, 'socketio_broadcast'):
        try:
            if event_type in app.socketio_broadcast:
                app.socketio_broadcast[event_type](data)
        except Exception as e:
            app.logger.error(f"SocketIO broadcast failed: {e}")

# Make helpers available in app context
app.trigger_integration = trigger_integration
app.broadcast_socketio = broadcast_socketio

if __name__ == '__main__':
    setup()

    # Run with SocketIO if available
    if SOCKETIO_AVAILABLE and socketio:
        app.logger.info("🚀 Starting InventoryApp v2.3.0 with SocketIO")
        socketio.run(app, host='0.0.0.0', port=5000, debug=app.debug)
    else:
        app.logger.info("🚀 Starting InventoryApp v2.3.0")
        app.run(host='0.0.0.0', port=5000)
