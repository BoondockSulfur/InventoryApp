# -*- coding: utf-8 -*-
"""
Advanced Reporting Engine for InventoryApp
Generates reports in multiple formats (PDF, Excel, CSV)
"""

import io
import csv
from datetime import datetime, timedelta
from flask import make_response

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not available - Excel export will be disabled")

try:
    from weasyprint import HTML
    PDF_AVAILABLE = True
except (ImportError, OSError):
    PDF_AVAILABLE = False
    print("⚠️  WeasyPrint not available - PDF export will be disabled")


class ReportGenerator:
    """Base class for report generation"""

    def __init__(self, db):
        self.db = db

    def generate_inventory_report(self, format='csv', filters=None):
        """
        Generate comprehensive inventory report

        Args:
            format: 'csv', 'excel', or 'pdf'
            filters: dict with filter criteria
        """
        from app import Item, Category, Location

        # Build query
        query = self.db.session.query(Item)

        if filters:
            if filters.get('category_id'):
                query = query.filter(Item.category_id == filters['category_id'])
            if filters.get('location_id'):
                query = query.filter(Item.location_id == filters['location_id'])
            if filters.get('is_borrowed') is not None:
                query = query.filter(Item.is_borrowed == filters['is_borrowed'])
            if filters.get('defective') is not None:
                query = query.filter(Item.defective == filters['defective'])

        items = query.all()

        if format == 'csv':
            return self._generate_csv_report(items)
        elif format == 'excel' and EXCEL_AVAILABLE:
            return self._generate_excel_report(items)
        elif format == 'pdf' and PDF_AVAILABLE:
            return self._generate_pdf_report(items)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_csv_report(self, items):
        """Generate CSV report"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            'ID', 'Name', 'Description', 'Serial', 'Category',
            'Location', 'Quantity', 'Borrowed', 'Defective', 'Created'
        ])

        # Data
        for item in items:
            writer.writerow([
                item.id,
                item.name,
                item.description or '',
                item.serial or '',
                item.category.name if item.category else '',
                item.location.name if item.location else '',
                item.quantity,
                'Yes' if item.is_borrowed else 'No',
                'Yes' if item.defective else 'No',
                item.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'created_at') and item.created_at else ''
            ])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        return response

    def _generate_excel_report(self, items):
        """Generate Excel report with formatting"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        headers = ['ID', 'Name', 'Description', 'Serial', 'Category',
                   'Location', 'Quantity', 'Borrowed', 'Defective', 'Created']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Data
        for item in items:
            row = [
                item.id,
                item.name,
                item.description or '',
                item.serial or '',
                item.category.name if item.category else '',
                item.location.name if item.location else '',
                item.quantity,
                'Yes' if item.is_borrowed else 'No',
                'Yes' if item.defective else 'No',
                item.created_at.strftime('%Y-%m-%d %H:%M') if hasattr(item, 'created_at') and item.created_at else ''
            ]
            ws.append(row)

            # Apply border to all cells
            for cell in ws[ws.max_row]:
                cell.border = border

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    def _generate_pdf_report(self, items):
        """Generate PDF report"""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #366092; border-bottom: 2px solid #366092; padding-bottom: 10px; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                th {{ background-color: #366092; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .footer {{ margin-top: 30px; text-align: center; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h1>Inventory Report</h1>
            <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p><strong>Total Items:</strong> {len(items)}</p>

            <table>
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>Name</th>
                        <th>Serial</th>
                        <th>Category</th>
                        <th>Location</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
        """

        for item in items:
            status = 'Defective' if item.defective else ('Borrowed' if item.is_borrowed else 'Available')
            html_content += f"""
                    <tr>
                        <td>{item.id}</td>
                        <td>{item.name}</td>
                        <td>{item.serial or 'N/A'}</td>
                        <td>{item.category.name if item.category else 'N/A'}</td>
                        <td>{item.location.name if item.location else 'N/A'}</td>
                        <td>{status}</td>
                    </tr>
            """

        html_content += """
                </tbody>
            </table>
            <div class="footer">
                <p>InventoryApp - Generated by Advanced Reporting Engine</p>
            </div>
        </body>
        </html>
        """

        pdf_bytes = HTML(string=html_content).write_pdf()

        response = make_response(pdf_bytes)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=inventory_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return response

    def generate_loan_report(self, format='csv', date_from=None, date_to=None):
        """Generate loan activity report"""
        from app import Loan, Item, User

        query = self.db.session.query(Loan)

        if date_from:
            query = query.filter(Loan.loan_date >= date_from)
        if date_to:
            query = query.filter(Loan.loan_date <= date_to)

        loans = query.all()

        if format == 'csv':
            return self._generate_loan_csv(loans)
        elif format == 'excel' and EXCEL_AVAILABLE:
            return self._generate_loan_excel(loans)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def _generate_loan_csv(self, loans):
        """Generate CSV loan report"""
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow([
            'ID', 'Item', 'Borrower', 'Loan Date', 'Due Date',
            'Return Date', 'Status'
        ])

        for loan in loans:
            status = 'Returned' if loan.return_date else (
                'Overdue' if loan.due_date < datetime.utcnow() else 'Active'
            )
            writer.writerow([
                loan.id,
                loan.item.name,
                loan.borrower.username,
                loan.loan_date.strftime('%Y-%m-%d'),
                loan.due_date.strftime('%Y-%m-%d'),
                loan.return_date.strftime('%Y-%m-%d') if loan.return_date else 'Not returned',
                status
            ])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=loan_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        return response

    def _generate_loan_excel(self, loans):
        """Generate Excel loan report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loan Report"

        # Header
        headers = ['ID', 'Item', 'Borrower', 'Loan Date', 'Due Date', 'Return Date', 'Status']
        ws.append(headers)

        # Styling
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Data
        for loan in loans:
            status = 'Returned' if loan.return_date else (
                'Overdue' if loan.due_date < datetime.utcnow() else 'Active'
            )
            ws.append([
                loan.id,
                loan.item.name,
                loan.borrower.username,
                loan.loan_date.strftime('%Y-%m-%d'),
                loan.due_date.strftime('%Y-%m-%d'),
                loan.return_date.strftime('%Y-%m-%d') if loan.return_date else 'Not returned',
                status
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=loan_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response


def get_report_generator(db):
    """Factory function to create report generator"""
    return ReportGenerator(db)
